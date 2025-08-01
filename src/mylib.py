#!/usr/bin/evn python3
#Copyright 2021, Zhengui WANG
#Apache License, Version 2.0; http://www.apache.org/licenses/LICENSE-2.0

from pylib import *

def ntype(data,fmt=0,mode='r',name=None,vshape=None,vtype=None):
    '''
    convert data to new self-defined datatype
    fmt=0: for numpy array in *.npz file
    fmt=1: for numpy array in *.nc file
    '''
    class narray: #derived type based on numpy.array
          def __init__(self,ndata):
              if fmt==1:
                 for i in data.ncattrs(): self.__dict__[i]=data.getncattr(i) #get all attribute
          def __getitem__(self,key): #subset of values
             data=self.value[key] if fmt==0 else self.data[key]; return array(data) if size(data)!=1 else data.item() if isinstance(data,np.ndarray) else data
          def __setitem__(self,key,item): #set variable
              if mode=='r':
                 if not hasattr(self,'_value'): self.value
                 self._value[key]=item
              elif mode in ['r+','w','w+','a']:
                  data[key]=item
              else:
                  sys.exit('simple fix: add open mode above')
          def __repr__(self):
              return 'array({})'.format(self.value)

          @property
          def value(self):   #all values of nc variable will be read out
              if not hasattr(self,'_value'): self._value=data[name] if fmt==0 else array(data[:])
              return self._value
          @property
          def data(self):
              return self._value if hasattr(self,'_value') else data

          @property
          def T(self):
              return self.value.T
          @property
          def ctypes(self):
              return self.value.ctypes
          @property
          def base(self):
              return self._value.base if hasattr(self,'_value') else data
          @property
          def shape(self):
              return vshape if fmt==0 else self.data.shape
          @property
          def size(self):
              return prod(vshape) if fmt==0 else self.data.size
          @property
          def dtype(self):
              return vtype if fmt==0 else self.data.dtype
          @property
          def flags(self):
              return self.data.flags
          @property
          def flat(self):
              return self.data.flat
          @property
          def imag(self):
              return self.data.imag
          @property
          def real(self):
              return self.data.real
          @property
          def ndim(self):
              return self.data.ndim
          @property
          def itemsize(self):
              return self.data.itemsize
          @property
          def nbytes(self):
              return self.data.nbytes
          @property
          def strides(self):
              return self.data.strides
          @property
          def INFO(self):
              return get_INFO(self)
          @property
          def VINFO(self):
              return get_INFO(self,1)

    def get_method(fn):
        def method(self, *args,**kwargs):
            return getattr(self.value, fn)(*args,**kwargs)
        return method
    blist=['shape','size','dtype', 'T','base','data','flags','flat','imag','real','ndim','itemsize','nbytes','strides','ctypes', #attributes re-written
           '__init__', '__repr__','__getitem__','__setitem__', #method re-written
           '__new__', '__getattribute__', '__getnewargs__', '__doc__', '__setattr__','__str__'] 
    datatype=type('narray',(narray,), {i: get_method(i) for i in dir(np.ndarray) if (i not in blist) and (not i.startswith('__array_'))})
    return datatype(data)

#def type(*args,**kwargs):
#    import builtins
#    return np.ndarray if (('narray' in str(builtins.type(args[0]))) and len(args)==1) else builtins.type(*args,**kwargs) 

def add_var(svar,value,sdict,fmt=0):
    '''
    add dictionary elements
      svar: variable name or list of variable names
      value: variable value or list of variable values
      sdict: dictionary
      fmt=0: ignore if variable already existing in sdict; fmt=1: always add: sdict[svar]=value
    '''
    if isinstance(svar,str): #single value
       if (svar not in sdict) or fmt==1: sdict[svar]=value
    else:  #list of values
       for i, k in zip(svar,value): add_var(i,k,sdict,fmt)

def sort_all(t,*args):
    '''
    sort all variables based on the 1st input
    '''
    sind=argsort(t); t=t[sind]; args=[i[sind] for i in args]
    return [t,*args]

def nargout():
   import traceback
   return len(traceback.extract_stack()[-3].line.split('=')[0].split(','))

def add_xtick(nts=6,xlim=None,xts=None,xls=None,grid='on',fmt='%Y-%m-%d\n%H:%M:%S',ax=None,fig=None):
    '''
    add dynamic xtick labels with x-axis being datenum
        nts:   number of xticklabel
        xlim:  x-axis limits
        xts:   xticks
        xls:   xticklables
        grid:  add x-axis grid line if grid='on'
        fmt:  datenum time string format; (=0:'%Y-%m-%d'; =1: '%d/%m')
        ax:    figure axes
        fig:   figure handle
    '''
    def update_xts(xm=None,xts=None,xls=None,tag=None):

        fstr= '%Y-%m-%d' if fmt==0 else '%d/%m' if fmt==1 else fmt
        if xm is None: xm=ax.get_xlim()
        if xts is None: xts=linspace(*xm,nts)
        if xls is None: xls=[num2date(i).strftime(fstr) for i in xts]
        ax.xaxis.set_ticks(xts); ax.xaxis.set_ticklabels(xls)
        ax.set_xlim(xm); ax.xaxis.grid(grid)
        if (tag==0 and fig.np==ax.np) or tag is None: fig.canvas.draw()

    def onclick(sp):
        dlk=int(sp.dblclick); btn=int(sp.button); x,y=sp.x,sp.y
        p=ax.get_window_extent(); x0,x1,y0,y1=p.x0,p.x1,p.y0,p.y1
        if x<x0 or x>x1 or y<y0 or y>y1: return
        if btn in [1,3]: update_xts()

    def init_xts(sp=None):
        update_xts(xm=xlim,xts=xts,xls=xls,tag=0)

    #pre-proc
    if fig is None: fig=gcf()
    if ax is None: ax=gca()
    #aviod canvas refresh from all axes
    fig.np=0 if not hasattr(fig,'np') else fig.np+1; ax.np=fig.np

    #set home
    if mpl.get_backend().lower() in ['qt5agg','qtagg']:
       acs=fig.canvas.toolbar.actions(); ats=[i.iconText().lower() for i in acs]
       ac=acs[ats.index('home')]; ac.triggered.connect(init_xts) #set home
    else: #work for tkagg
       try:
          ac=fig.canvas.toolbar.children['!button']
          ac.bind("<Button-1>", init_xts)
       except:
         pass

    #set init. xtick
    if xlim is None:
        xms=array([[i.get_xdata().min(),i.get_xdata().max()] for i in ax.get_children() if type(i)==mpl.lines.Line2D])
        xlim=[xms[:,0].min(),xms[:,1].max()] if xms.ndim==2 else ax.get_xlim()
    if xts is not None: nts=sum((array(xts)>=xlim[0])*(array(xts)<=xlim[1]))
    init_xts()

    #connect actions
    fig.canvas.mpl_connect('button_release_event', onclick)

def cmean(mtime,data,n=13):
    '''
    compute climatological pattern by using running smooth method for computing
       mtime, data: time and data
       n: number of points in a year for the outputs
    '''
    t0=doy(mtime) if mtime.max()>365 else mtime; y0=data; t=linspace(0,365,n); dt=t[1]-t[0]
    fmean=lambda x: nan if len(x)==0 else x.mean()
    y=array([fmean(y0[abs(t0-i)<=dt/2]) for i in t]); fpn=~isnan(y); t,y=t[fpn],y[fpn]
    return t,y

def urlsave(url,fname):
    '''
    download url to local, using urllib.request.urlretrieve
    '''
    import urllib,ssl
    try:
           _create_unverified_https_context = ssl._create_unverified_context
    except AttributeError: # Legacy Python that doesn't verify HTTPS certificates by default
           pass
    else: # Handle target environment that doesn't support HTTPS verification
           ssl._create_default_https_context = _create_unverified_https_context
    urllib.request.urlretrieve(url,fname)

def interp(x0,y0,x,fmt=0,kind='linear',axis=-1,fill_value=nan):
    '''
    1D interpolator with different options for outliers; The engine in scipy.interpolate.interp1d.
      x0,y0: orignal data
      x: coordindate that data will be evaluated
      kind: 'linear', 'linear', 'nearest', 'nearest-up', 'zero', 'slinear', 'quadratic', 'cubic', 'previous', 'next'
      axis: interpolation performed along certain dimension of y0

    Options for points outside the range of x0
      fmt=0: use nearest point
      fmt=1: simple linear extrapolate
      fmt=2: (x0,x) represents time; use nearest year for outlier interpolation
      fmt=3: (x0,x) represents time; compute climatological pattern for outlier interpolation

      note: fmt=3 only works for 1D data
    '''
    f=sp.interpolate.interp1d; x1=x0.min(); x2=x0.max(); x=array(x).copy().astype('float'); t0=x0; t1=x1; t2=x2; t=x

    #different options for points out of range
    if fmt==0: x[x<x1]=x1; x[x>x2]=x2
    if fmt==1: fill_value='extrapolate'
    if fmt==2:
       d1=doy(t1); d2=doy(t2); t10=t1-d1; t20=t2-d2 #minimum and maximum dates
       fpt=t<t1; d=doy(t[fpt]); tm=t10+d; fp=d<d1; tm[fp]=tm[fp]+365; t[fpt]=tm
       fpt=t>t2; d=doy(t[fpt]); tm=t20+d; fp=d>d2; tm[fp]=tm[fp]-365; t[fpt]=tm
    if fmt==3:
       fp=(t>=t1)*(t<=t2); fpn=~fp; ym=nan*ones(len(t))
       ym[fp]=f(x0,y0,kind=kind,axis=axis,fill_value=fill_value)(t[fp])
       if sum(fpn)!=0:
          ct,cy=cmean(t0,y0,int(t0.size/10)); npt=ct.size; ct=r_[ct-365,ct,ct+365]; cy=r_[cy,cy,cy]
          if npt>13: cy=smooth(cy,2+int(npt/13))
          ym[fpn]=interp(ct,cy,doy(t[fpn]))
    #interp
    y=ym if fmt==3 else f(x0,y0,kind=kind,axis=axis,fill_value=fill_value)(x)
    return y

class blit_manager:
    def __init__(self,hgs,hf=None):
        '''
        animation manager using blit function in matplotlib
        hgs: list of plot artists
        hf: figure handle

        Example:
            xi=arange(0,5*pi,0.1); yi=sin(xi)
            [hg]=plot(xi,yi,animated=True); bm=blit_manager(hg)
            for i in arange(100): hg.set_ydata(sin(xi-i)); bm.update()
        '''
        if hf is None: hf=gcf()
        if not isinstance(hgs,list): hgs=[hgs]

        pause(0.001); [i.set_animated(True) for i in hgs]
        self.hf=hf
        self.hgs=hgs
        self.update_bg(0)
        self.cid=self.hf.canvas.mpl_connect("draw_event", self.update_bg)

    def update_bg(self,event):
        self.bg=self.hf.canvas.copy_from_bbox(self.hf.bbox)
        for hg in self.hgs: self.hf.draw_artist(hg)

    def update(self):
        self.hf.canvas.restore_region(self.bg)
        for hg in self.hgs: self.hf.draw_artist(hg)
        self.hf.canvas.blit(self.hf.bbox)
        self.hf.canvas.flush_events()

def add_basemap(xm=None,ym=None,fmt=0,dpi=300,service='World_Imagery',resolution='h',prj='epsg:4326',scale=1,color='0.8',**args):
    '''
    add basemaps in Lon&Lat projection, and it needs packages of "basemap" and "basemap-data-hires"
    For more information about Basemap, refer to https://matplotlib.org/basemap/stable/users/geography.html

    xm: longitude range;  ym: latitude range
    prj: only a few projecitons are supported (see Basemap doc)

    fmt=0: arcgis map service
       dpi: resolution for arcgis map
       service: arcgis service (World_Imagery, World_Topo_Map, World_Street_Map, USA_Topo_Maps,etc.)
                (https://doc.arcgis.com/en/data-appliance/7.2/maps/world-topographic-map.htm)

    fmt>1: other basemaps (works better for larger region)
       fmt=1: costal line;  2: rivers;  3: lands;  4: US states;  5: etopo;   6: shadedrelief;  7:bluemarble
       resolution: (c,l,i,h,f) for other maps
       scale: downsample parameter for some basemaps
       color: color for fillcontinents
    '''
    #pre-proc
    try:
      from mpl_toolkits.basemap import Basemap
    except:
      sys.exit('install basemap')
    if xm is None: xm=xlim()
    if ym is None: ym=ylim()
    if not hasattr(fmt,'__len__'): fmt=[fmt]

    #get basemap
    bm=Basemap(llcrnrlon=xm[0],urcrnrlon=xm[1],llcrnrlat=ym[0],urcrnrlat=ym[1],epsg=prj.split(':')[-1],resolution=resolution,**args)
    for i in fmt:
        if i==0: bm.arcgisimage(xpixels=dpi,service=service)
        if i==1: bm.drawcoastlines()
        if i==2: bm.drawrivers()
        if i==3: bm.fillcontinents(color=color)
        if i==4: bm.drawstates()
        if i==5: bm.etopo(scale=scale)
        if i==6: bm.shadedrelief(scale=scale)
        if i==7: bm.bluemarble(scale=scale)
    return bm

class _COMM_WORLD:
      def __init__(self):
          self.pid=os.getpid()
          self.myrank=None

      def Get_size(self):
          self.nproc=1 if os.getenv('nproc') is None else int(os.getenv('nproc'))
          return self.nproc

      def Get_rank(self):
          header='.pylib_pjob_id_'; open(header+str(self.pid),'w+').close()
          if not hasattr(self,'nproc'): self.Get_size()
          while len(glob(header+'*'))<self.nproc: time.sleep(0.1)
          fnames=glob(header+'*'); pids=[*sort([int(i.split('_')[-1]) for i in fnames])]
          self.myrank=pids.index(self.pid); self.delete(header,0.2)
          return self.myrank

      def Barrier(self):
          header='.pylib_pjob_lock_'; open(header+str(self.pid),'w+').close()
          while len(glob(header+'*'))<self.nproc: time.sleep(0.1)
          self.delete(header,dt=0.2)

      def delete(self,header=None,dt=0):
          if self.myrank==0:
             time.sleep(dt); [os.remove(i) for i in glob(header+'*')]
          else:
             time.sleep(2*dt)

class parallel_jobs:
      '''
      compute nproc and myrank by dumping pid information on disk, when mpi4py fails.
      Note: this is not MPI, but only for the purpose in distributing parallel jobs
      '''
      def __init__(self):
          self.COMM_WORLD=_COMM_WORLD()

def resize(data,shape,fill=0):
    '''
    re-define the shape of an data array; The added data are filled with value='fill argument'
    '''
    ds=data.shape; fdata=zeros(shape).astype(data.dtype); fdata[:]=fill
    exec('fdata[{}]=data'.format(':'+',:'.join([str(i) for i in ds])+',0'*(len(shape)-len(ds))))
    return fdata

def mklink(fname,fmt=0):
    '''
      execute in windows: convert symbolic links of directory  to mklink (DOS)
      fname: directory name(s)
      fmt=0: skip symbolic sub-direcotries; fmt=1: include search in sub-directories
    '''

    if platform.system().lower()=="windows":
       #directory need to check symbolic links
       if isinstance(fname,str):
          sdirs=[fname]
       else:
          sdirs=[*fname]

       #change symlink recursively
       while len(sdirs)!=0:
           sdirs0=unique(array([os.path.abspath(i) for i in sdirs])); sdirs=[]
           for sdir in sdirs0:
               os.chdir(sdir); fnames=os.listdir(sdir)

               #change symlink
               for fname in fnames:
                   if os.path.isdir(fname): sdirs.append('{}/{}'.format(sdir,fname)); continue #gather directory
                   if not os.path.islink(fname): continue

                   tname=os.readlink(fname)
                   if not os.path.isdir(tname): continue
                   os.remove(fname); os.system('mklink /D {} {}'.format(fname,tname))
                   sdirs.append('{}/{}'.format(sdir,fname))
           if fmt==0: sdirs=[i for i in sdirs if not os.path.islink(i)]

def rtext(x,y,note,xm=None,ym=None,ax=None,**args):
    '''
    add annoation to the current axes at relative location to x-axis and y-axis
       x: relative location in x-axis, tpycially 0-1, but can be negative or larger than 1 
       y: relative location in y-axis, typically 0-1, but can be negative or larger than 1
       note: annotation string
       xm: range of x-axis; gca().xlim() is used if not given
       ym: range of y-axis; gca().ylim() is used if not given
       ax: axes; gca() is used if not given
       **args: all other parameters used in matploblib.plot.text are also applicable

    E.q., rtext(0.1,0.9,"(a)",size=10,weight='bold') -> place label at top left corner

    note: suggest not to rearrange the x/ylim after using this function, or place this after
    '''
    if ax is None: ax=gca()
    sca(ax)
    if xm is None: xm=xlim()
    if ym is None: ym=ylim()
    text(xm[0]+x*diff(xm),ym[0]+y*diff(ym)*y,note,**args)
    
def read_excel(fname,sht=0,fmt=0):
    '''
    use pd.read_excel to read Excel file
      fname: name of Excel file
      sht:   name of sheet_name, or number of sheet (default is 0)
      fmt=0: not treat 1st line as header, return data only; fmt=1: treat 1st line as header, return header and data
    '''
    import pandas as pd

    if fmt==0:
       fd=pd.read_excel(fname,sheet_name=sht,header=None)
    else:
       fd=pd.read_excel(fname,sheet_name=sht)
    header=array(fd.columns); fdata=fd.values

    if fmt==0: return fdata
    if fmt==1: return header, fdata

def write_excel(fname,data,sht='sheet_1',indy=0,indx=0,fmt=0,align='row',old_font=0,
                color=None,fontsize=None,fontweight=None,number_format=None,figsize=None,
                cell_width=None,cell_color=None,  **args):
    '''
    use xlsxwriter to write Excel file
       fname: name of Excel file
       data: can be a single data, 1D array or 2D array
       sht:  name of sheet_name
       indy: starting Row index of cell for data to be written
       indx: starting Column index of cell for data to be written
       align='row': write 1D data as a row; align='column': write 1D data as a column
       old_font=1: for existing sheet,keep old font styles; old_font=0: discard
       color,fontsize,fontweight: options for specify cell font (see openpyxl.styles.Font)
       number_format: specifiy the data format (e.g. "0.0000")
       cell_width: cell width
       cell_color: cell color in the background

       fmt=0: append data to existing file, or create file if not existing
       fmt=1: replace mode of excel file
       fmt=2: only replace sheet, but keep other sheets of excel file
       fmt=3: insert image (data is a figure object or path of an image, and figsize=[w,h] is used to resize the figure in pixels)
    '''
    try: 
       import openpyxl,io
       import pandas as pd
       #import xlsxwriter as xw
    except:
       sys.exit('install openpyxl pandas')

    #open fname
    if not fname.endswith('.xlsx'): fname=fname+'.xlsx'
    if os.path.exists(fname) and (fmt in [0,2,3]):
        fid=pd.ExcelWriter(fname,mode='a',engine='openpyxl',if_sheet_exists='overlay')
    else:
        fid=pd.ExcelWriter(fname,mode='w+',engine='openpyxl')

    if fmt in [0,1,2]: #write cell data
       #reorganize data to a 2D array
       if isinstance(data,str) or (not hasattr(data,'__len__')): data=array([[data,],]) #for single data
       data=array(data) #for list
       if data.ndim==1: data=data[None,:] if (align=='row') else data[:,None]  #for 1D array

       #get dimension info
       ds=data.shape; ny, nx=ds[0]+indy, ds[1]+indx

       #create matrix of fields
       if (sht in list(fid.sheets)) and fmt==0:
           data0=read_excel(fname,sht).astype('O'); ny0,nx0=data0.shape
           if old_font==1:
              sid=fid.sheets[sht]; fonts=ones([ny0,nx0]).astype('O')
              for k in arange(ny0):
                  for i in arange(nx0): fonts[k,i]=sid.cell(k+1,i+1).font.copy()
           if ny0<ny: data0=r_[data0,tile('',[(ny-ny0),nx0])]; ny0=ny
           if nx0<nx: data0=c_[data0,tile('',[ny0,(nx-nx0)])]; nx0=nx
       else:
           data0=tile('',[ny,nx]).astype('O')

       #write data
       for k, datai in enumerate(data):
           for i, dataii in enumerate(datai): data0[indy+k,indx+i]=dataii

       df=pd.DataFrame([list(i) for i in data0])
       df.to_excel(fid,sheet_name=sht,header=False,index=False)

       #specify cell font
       if old_font==1 and (sht in list(fid.sheets)) and fmt==0:
          sid=fid.sheets[sht]
          for k in arange(fonts.shape[0]):
              for i in arange(fonts.shape[1]): sid.cell(k+1,i+1).font=fonts[k,i]

       #cell format
       sid=fid.sheets[sht]
       if color is not None: color=mpl.colors.to_hex(color)[1:]
       for k, datai in enumerate(data):
           for i, dataii in enumerate(datai):
               if cell_color is not None:
                  clr=mpl.colors.to_hex(cell_color).upper()[1:]
                  sid.cell(indy+k+1,indx+i+1).fill=openpyxl.styles.PatternFill(start_color=clr, end_color=clr, fill_type="solid")
               if (color,fontsize,fontweight)!=(None,None,None) or len(args)!=0:
                  #from openpyxl.styles import Color, PatternFill, Font, Border
                  cf=openpyxl.styles.Font(color=color,size=fontsize,bold=(fontweight=='bold'),**args)
                  sid.cell(indy+k+1,indx+i+1).font=cf
               if number_format is not None: sid.cell(indy+k+1,indx+i+1).number_format=number_format
               if k==0 and cell_width is not None: sid.column_dimensions[sid.cell(indy+k+1,indx+i+1).column_letter].width=cell_width
                  
    elif fmt==3: #insert image
      if isinstance(data,mpl.figure.Figure):
         buf=io.BytesIO(); data.savefig(buf,format='png'); buf.seek(0); idata=openpyxl.drawing.image.Image(buf)
      else:
         idata=openpyxl.drawing.image.Image(data)
      sid=fid.sheets[sht]
      if figsize is not None: idata.width=figsize[0]; idata.height=figsize[1]
      sid.add_image(idata,sid.cell(indy,indx).coordinate)
    fid.close()

def read_yaml(fname):
    '''
    read yaml file and return key-value dict
    
    '''
    lines=open(fname,'r').readlines()
    param={}
    for line in lines:
        sline=line.strip().split(':')
        if len(sline)<2: continue
        key=sline[0].strip(); value=sline[1].strip()
        if value=='': continue
        value=value.split()[0]
        param[key]=value

    return param

def get_qnode(qnode=None):
    '''
    return hpc node name based on system's environment variable $HOST
    '''
    import re
    if qnode is not None: return qnode  #used other node names

    host=os.getenv('HOST')
    if host is None: sys.exit('set HOST environment variable HOST')
    if host in ['femto.sciclone.wm.edu','viz']: qnode='femto'
    if host in ['kuro.sciclone.wm.edu']: qnode='kuro'
    if host in ['gulf.sciclone.wm.edu']: qnode='gulf'
    if re.match('bo\w+.sciclone.wm.edu',host)!=None: qnode='bora'
    if host.startswith('bo') and host.endswith('sciclone.wm.edu'): qnode='bora'
    if host in ['chesapeake.sciclone.wm.edu']: qnode='potomac'
    return qnode

def get_hpc_command(code,bdir,jname='mpi4py',qnode=None,nnode=1,ppn=1,wtime='01:00:00',scrout='screen.out',
                    fmt=0,ename='param',qname=None,account=None,reservation=None,mem=None,exclusive=0):
    '''
    get command for batch jobs on sciclone/ches/viz3
       code: job script
       bdir: current working directory
       jname: job name
       qname: partition name (needed on some cluster/project)
       qnode: hpc node name
       nnode,ppn,wtime: request node number, core per node, and walltime
       mem: meomory; needed on cluster grace
       reservation: HPC reservation information
       fmt=0: command for submitting batch jobs; fmt=1: command for run parallel jobs
    '''

    #pre-proc
    qnode0=qnode; qnode=os.getenv('qnode') if qnode is None else qnode; nproc=nnode*ppn
    if ename=='run_schism': ename='schism' #old parameter support

    #combine commands for slurm system (abandon pbs system)
    jname='-J '+jname; nnode='-N {}'.format(nnode);  ppn='--ntasks-per-node {}'.format(ppn); wtime='-t '+wtime #necessary
    qname='' if qname==None else '-p '+qname
    account='' if account==None else '-A '+account
    reservation='' if reservation==None else '--reservation='+reservation
    mem='' if mem==None else '--mem='+mem
    exclusive='' if exclusive==0 else '--exclusive' #'levante','hercules'

    if fmt==0:
       os.environ[ename]='{} {}'.format(bdir,os.path.abspath(code)); os.environ['run_schism']=os.getenv(ename)
       #for submit jobs
       if qnode in ['femto','gulf','kuro','bora','frontera','levante','hercules','eagle','deception','grace','stampede2','stampede3']:
          scmd='sbatch --export=ALL {} {} {} {} {} {} {} {} {} {}'.format(qname,account,reservation,mem,exclusive,jname,nnode,ppn,wtime,code)
       else:
          sys.exit('unknown qnode: {},tag=1'.format(qnode))
       #if qnode in ['stampede2',]: scmd='sbatch "--export=ALL" {} {} {} {} -n {} {} {}'.format(jname,qname,account,nnode,nproc,wtime,code)
    elif fmt==1:
       #for run parallel jobs
       ALL='ALL,' if ename=='schism' else '' #ALL='' if os.getenv('mpi_env')=='openmpi' else 'ALL,'
       if qnode in ['femto','gulf','kuro','bora','stampede3']:
          scmd='srun --export={}PATH={},LD_LIBRARY_PATH={},job_on_node=1,bdir={},nproc={} {} >& {}'.format(ALL,os.path.dirname(sys.executable),os.getenv('LD_LIBRARY_PATH'),bdir,nproc,code,scrout)
          #if ename=='schism': scmd='srun --export=ALL,job_on_node=1,bdir={},nproc={} {} >& {}'.format(bdir,nproc,code,scrout)
       elif qnode in ['frontera']:
          scmd="mpirun --env job_on_node 1 --env bdir='{}' --env nproc {} -np {} {} >& {}".format(bdir,nproc,nproc,code,scrout)
          if ename=='schism': scmd="ibrun {} >& {}".format(code,scrout)
       elif qnode in ['stampede2',]:
          scmd="mpiexec -envall -genv job_on_node 1 -genv bdir '{}' -genv nproc {} -n {} {} >& {}".format(bdir,nproc,nproc,code,scrout)
          if ename=='schism': scmd="ibrun {} >& {}".format(code,scrout)
       elif qnode in ['levante','hercules']:
          scmd="mpiexec -envall -genv job_on_node 1 -genv bdir '{}' -genv nproc {} -n {} {} >& {}".format(bdir,nproc,nproc,code,scrout)
          if ename=='schism':
             scmd="ulimit -s unlimited; ulimit -c 0; source /home/g/g260135/intel_tool; export UCX_UNIFIED_MODE=y;"
             scmd=scmd+"srun --export=ALL,job_on_node=1,bdir={} -l --cpu_bind=verbose --hint=nomultithread --distribution=block:cyclic {} >& {}".format(bdir,code,scrout)
          if qnode=='hercules' and ename=='schism':
             scmd="set -e; ulimit -s unlimited; source /home/yjzhang/modules.hercules;"
             scmd=scmd+"srun --export=ALL,job_on_node=1,bdir={},nproc={} {} >& {}".format(bdir,nproc,code,scrout)
       elif qnode in ['eagle','deception']:
          scmd="mpirun --env job_on_node 1 --env bdir='{}' -{} {} {} >& {}".format(bdir,'n' if qnode=='eagle' else 'np',nproc,code,scrout)
       elif qnode in ['grace',]:
          scmd="mpirun --env job_on_node 1 --env bdir='{}' -np {} {} >& {}".format(bdir,nproc,code,scrout) 
          if ename=='schism': scmd="mpirun -np {} ./{} >& {}".format(nproc,code,scrout)
       else:
          sys.exit('unknown qnode: {},tag=2'.format(qnode))

    return scmd

def compute_contour(x,y,z,levels,fname=None,prj='epsg:4326',show_contour=False,nx=5000,ny=5000):
    '''
    compute contour lines
    Input:
        x: array for x coordinates (ndx)
        y: array for y coordinates (ndy)
        z: matrix of data (ndy,ndx)
        levels: values of contour lines to be extracted
        (nx,ny): when ndx(ndy)>nx(ny), subdivide the domain to speed up
    Output:
        fname: when fname is not None, write contours in shapefiles
        prj:  projection names of shapefiles
        show_contour: plot contours
    '''

    #check level
    fpn=~isnan(z); zmin=z[fpn].min(); zmax=z[fpn].max()
    levels=array(levels); fpz=(levels>=zmin)*(levels<=zmax); levels=sort(levels[fpz])

    #data capsule
    S=zdata(); S.levels=levels; S.xy=[[] for i in arange(len(levels))]
    if len(levels)==0: return S

    #divide domain
    ndx=len(x); ndy=len(y)
    ixs=[]; i1=0; i2=min([nx,ndx])
    while True:
        ixs.append([i1,i2])
        if i2>=ndx: break
        i1=i1+nx; i2=min([i2+nx,ndx])

    iys=[]; i1=0; i2=min([ny,ndy])
    while True:
        iys.append([i1,i2])
        if i2>=ndy: break
        i1=i1+ny; i2=min([i2+ny,ndy])

    #extract contours for subdomains
    for m,[ix1,ix2] in enumerate(ixs):
        for n,[iy1,iy2] in enumerate(iys):
            sxi=x[ix1:ix2]; syi=y[iy1:iy2]; szi=z[iy1:iy2,ix1:ix2]
            fpn=~isnan(szi); zmin=szi[fpn].min(); zmax=szi[fpn].max()
            fpz=(levels>=zmin)*(levels<=zmax); levels_sub=sort(levels[fpz])
            if len(levels_sub)==0: continue
            print('extracting contours in subdomain: {}/{}'.format(n+1+m*len(iys),len(ixs)*len(iys)))

            hf=figure(visible=False); P=contour(sxi,syi,szi,levels_sub); close(hf)
            if 'get_paths' in P.__dir__():
               for k,p in enumerate(P.get_paths()):
                   pxy=[]; [pxy.extend(r_[pn,nan*ones([1,2])]) for pn in p.to_polygons(closed_only=False)] #for each level
                   sindc=pindex(levels,levels_sub[k])[0]; S.xy[sindc].extend(pxy) #collect contour in each subdomain
            else: #old mpl version
               for k,p in enumerate(P.collections):
                   pxy=[]; [pxy.extend(r_[pn.vertices,nan*ones([1,2])]) for pn in p.get_paths()] #for each level
                   sindc=pindex(levels,levels_sub[k])[0]; S.xy[sindc].extend(pxy) #collect contour in each subdomain
    for i in arange(len(levels)): S.xy[i]=array(S.xy[i])

    #write contours
    if fname is not None:
        for i,vi in enumerate(levels):
            c=zdata(); c.type='POLYLINE'; c.xy=S.xy[i]; c.prj=get_prj_file(prj)
            cname='{}_{}'.format(fname,vi) if vi>0 else '{}_m{}'.format(fname,-vi)
            write_shapefile_data(cname,c)
    #plot contours
    if show_contour:
        figure(); cs='krgbmcy'
        for i,vi in enumerate(levels): plot(*S.xy[i].T,color=cs[i%7],lw=0.5)
        legend([*levels])
    return S

def load_dem(x,y,fname,z=None,fmt=0,position='center'):
    '''
    load bathymetry data onto points(xy)
    Input:
        fname: name of DEM file (format can be *.asc, *.npz, *.tiff or schism grid "grid.npz" )

    Outpt:
        fmt=0: return dp; depth interpolated for all points
        fmt=1: return [dpi, sindi]; depth for points only modified, and also flag
    '''

    #input
    xi0=x.copy(); yi0=y.copy(); igrd=0

    #read DEM lon, lat
    if fname.endswith('npz'):
        try:
          S=loadz(fname,['ne','np']); igrd=1 #unstructured schism grid
        except:
          S=loadz(fname,['lon','lat']); dx=abs(diff(S.lon)).mean(); dy=abs(diff(S.lat)).mean()
    elif fname.endswith('asc')or fname.endswith('.tif') or fname.endswith('.tiff'):
        S=read_dem(fname,fmt=1,position=position); dx=abs(diff(S.lon)).mean(); dy=abs(diff(S.lat)).mean()
    else:
        sys.exit('wrong format of DEM')

    #check domain of DEM
    if igrd==0:
       xm=[S.lon.min()-dx/2, S.lon.max()+dx/2]; ym=[S.lat.min()-dy/2, S.lat.max()+dy/2]
    else:
       from .schism_file import schism_grid
       gd=schism_grid(); gd.__dict__=loadz(fname).__dict__
       xm=[gd.x.min(),gd.x.max()]; ym=[gd.y.min(),gd.y.max()]
    if xi0.min()>=xm[1] or xi0.max()<=xm[0] or yi0.min()>=ym[1] or yi0.max()<=ym[0]: #return depth
       if fmt==0:
          if z is None: z=zeros(len(x))*nan
          return z
       elif fmt==1:
          return [array([]),array([]).astype('int')]
       else:
          sys.exit('wrong fmt')

    if igrd==0: #raster file
       #load DEM data
       if fname.endswith('npz'):
           S=loadz(fname)
       else:
           S=read_dem(fname,position=position)
       if not hasattr(S,'nodata'): S.nodata=None

       #change y direction
       if mean(diff(S.lat))<0: S.lat=flipud(S.lat); S.elev=flipud(S.elev)
       lon=S.lon; dx=diff(lon).mean()
       lat=S.lat; dy=diff(lat).mean()
       lon1=lon.min(); lon2=lon.max(); lat1=lat.min(); lat2=lat.max()

       #move (x,y) by half cell
       fpn=(xi0>=(lon1-dx/2))*(xi0<lon1); xi0[fpn]=lon1
       fpn=(yi0>=(lat1-dy/2))*(yi0<lat1); yi0[fpn]=lat1
       fpn=(xi0>=lon2)*(xi0<=(lon2+dx/2)); xi0[fpn]=lon2-dx*1e-6
       fpn=(yi0>=lat2)*(yi0<=(lat2+dy/2)); yi0[fpn]=lat2-dy*1e-6

       #get (x,y) inside dem domain
       sindp=pindex((xi0>=lon1)*(xi0<=lon2)*(yi0>=lat1)*(yi0<=lat2))
       xi=xi0[sindp]; yi=yi0[sindp]

       #compute index of (x,y)
       idx=floor((xi-lon[0])/dx).astype('int')
       idy=floor((yi-lat[0])/dy).astype('int')

       #make sure lon[idx]<=xi
       sind=pindex((lon[idx]-xi)>0)
       while len(sind)!=0:
           idx[sind]=idx[sind]-1
           fps=pindex((lon[idx[sind]]-xi[sind])>0)
           sind=sind[fps]

       #make sure lat[idy]<=yi
       sind=pindex((lat[idy]-yi)>0)
       while len(sind)!=0:
           idy[sind]=idy[sind]-1
           fps=pindex((lat[idy[sind]]-yi[sind])>0)
           sind=sind[fps]

       #compute xrat and yrat
       xrat=(xi-lon[idx])/(lon[idx+1]-S.lon[idx])
       yrat=(yi-lat[idy])/(lat[idy+1]-S.lat[idy])
       if sum((xrat<0)*(xrat>1))!=0: sys.exit('xrat<0 or xrat>1')
       if sum((yrat<0)*(yrat>1))!=0: sys.exit('yrat<0 or yrat>1')

       #make sure elevation is within right range
       if S.nodata is not None:
           if isnan(S.nodata):
              fpz=isnan(S.elev)
           else:
              #fpz=(abs(S.elev)>1.5e4)|(S.elev==S.nodata)|(abs(S.elev-S.nodata)<1)
              fpz=S.elev==S.nodata
           S.elev[fpz]=nan

       #compute depth
       dp1=S.elev[idy,idx]*(1-xrat)+S.elev[idy,idx+1]*xrat
       dp2=S.elev[idy+1,idx]*(1-xrat)+S.elev[idy+1,idx+1]*xrat
       dp=dp1*(1-yrat)+dp2*yrat

       #make sure elevation is within right range
       if sum(isnan(S.elev))>0:
           fpz=~isnan(dp); sindp=sindp[fpz]; dp=dp[fpz]

    elif igrd==1: #schism grid (unstrucured data)
       sind=pindex((xi0>=xm[0])*(xi0<=xm[1])*(yi0>=ym[0])*(yi0<=ym[1]))
       pie,pip,pacor=gd.compute_acor(c_[xi0[sind],yi0[sind]],out=0); fp=pie!=-1
       if sum(fp)!=0:
          sindp=sind[fp]; dp=-(gd.dp[pip[fp]]*pacor[fp]).sum(axis=1)
       else:
          sindp=array([]).astype('int'); dp=array([])

    #return depth
    if fmt==0:
        if z is None: z=zeros(len(x))*nan
        z[sindp]=dp
        return z
    elif fmt==1:
        return [dp, sindp]
    else:
        sys.exit('wrong fmt')

def rewrite_input(fname,qnode=None,nnode=1,ppn=1,**args):
    '''
    function to rewrite the inputs in job-submit scripts (e.g.run_mpi_template.py)
    '''
    #write qnode,nnode,ppn
    if qnode is None: sys.exit('please specify qnode')
    rewrite(fname,replace=['qnode','#qnode'],startswith=['qnode='])
    rewrite(fname,replace=["qnode='{}'; nnode={}; ppn={}\n".format(qnode,nnode,ppn)],startswith=["#qnode='{}'".format(qnode)],note_delimiter='#')       

    #change parameters
    for key,value in args.items(): 
        if key in ['icmb','ifs','fmt','stacks']: 
           fstr="{}={}".format(key,value)
        else:
           fstr="{}='{}'".format(key,value)
        rewrite(fname,replace=[fstr],startswith=['{}='.format(key)],note_delimiter='#')

def rewrite(fname,fmt=0,replace=None,include=None,startswith=None,endswith=None,append=None,note_delimiter=None):
    '''
    function to rewrite file in-situ based on conditions
         fname: file name
         fmt: remove whitespace at line head and tail (0: not; 1: head & tail; 2: head only;  3: tail only)
         replace: string pairs list; e.g. replace=['ON', 'OFF']
         include: list of strings included; e.g. include=['USE_ICM']
         startswith: list of  strings startswith; e.g. startswith=['#USE_ICM']
         endswith: list of strings endswith;   e.g. endwith=['*.csv']
         append: list of lines; e.g. ['add 1st line','add 2nd line']
         note_delimiter: keep inline note after delimiter (ignore delimiters in the beginning)
    '''

    if isinstance(replace,str): replace=[replace]
    if isinstance(include,str): include=[include]
    if isinstance(startswith,str): startswith=[startswith]
    if isinstance(endswith,str): endswith=[endswith]

    #read fname
    if os.path.exists(fname):
       fid=open(fname,'r'); lines=fid.readlines(); fid.close()
    else:
       return

    #rewrite lines
    slines=[]
    for line in lines:
        sline=line; iflag=0

        #whitespace in the begining and end
        if fmt==1: sline=sline.strip()
        if fmt==2: sline=sline.lstrip()
        if fmt==3: sline=sline.rstrip()

        #check include
        if include is not None:
           for i in include:
               if i in sline: iflag=1

        #check startswith
        if startswith is not None:
           for i in startswith:
               if sline.strip().startswith(i): iflag=1

        #check startswith
        if endswith is not None:
           for i in endswith:
               if sline.strip().endswith(i): iflag=1

        #check note
        nd=note_delimiter; note=''
        if (nd is not None) and iflag==1:
           note=sline.strip()
           while note.startswith(nd): note=note[1:]
           if nd in note:
              sid=note.find(nd); note=note[sid:]
           else:
              note=''

        #replace string
        if iflag==1:
           if replace is not None:
              if len(replace)==0:
                 continue
              elif len(replace)==1:
                 sline=replace[0].rstrip()+' '+note
              else:
                 sline=sline.replace(*replace)
           else:
              continue

        #save new line
        if not sline.endswith('\n'): sline=sline+'\n'
        slines.append(sline)

    #append
    if append is not None:
       for sline in append:
           if not sline.endswith('\n'): sline=sline+'\n'
           slines.append(sline)

    #write new line
    fid=open(fname,'w+'); fid.writelines(slines); fid.close()

def read_dem(fname,sname=None,fmt=0,position='center'):
    '''
    fname: name of source DEM (*.asc, *.tif, *.tiff) file
    sname (optional): name of file to be saved if value is provided 
    fmt=0: read all the DEM data
    fmt=1: only read domain information (lon, lat)
    fmt=2: return geotiff_metadata

    Note: for *.tif and *.tiff files:
      position='center': elevation is defined at cell center
      position='corner': elevation is defined at cell corner; coordinates will be shifted by half a cell.
    '''

    #read dem data
    if fname.endswith('.tif') or fname.endswith('.tiff'):
       try:
          import tifffile as tiff
       except:
          sys.exit('install tifffile=2022.5.4')
       ginfo=tiff.TiffFile(fname).geotiff_metadata; S=zdata()
       if fmt==2: return ginfo
       dx,dy=ginfo['ModelPixelScale'][:2]; xll,yll=ginfo['ModelTiepoint'][3:5]
       if position=='corner': xll=xll+dx/2; yll=yll-dy/2
       try:
          from PIL import Image
          sinfo=Image.open(fname)
          nrows=sinfo.height; ncols=sinfo.width
       except:
          nrows,ncols=tiff.imread(fname).shape
       lon=xll+dx*arange(ncols); lat=yll-dy*arange(nrows)
       if fmt==0: elev=tiff.imread(fname).astype('float32'); elev[abs(elev)>=9999]=-9999.0; S.elev=elev
       S.lon=lon; S.lat=lat; S.nodata=-9999.0
    else: #must be *.asc format
        if not fname.endswith('.asc'): fname=fname+'.asc'
        #read file
        fid=open(fname,'r'); S=zdata()
        ncols=int(fid.readline().strip().split()[1])
        nrows=int(fid.readline().strip().split()[1])
        xn,xll=fid.readline().strip().split(); xll=float(xll)
        yn,yll=fid.readline().strip().split(); yll=float(yll)
        dxy=float(fid.readline().strip().split()[1])
        S.nodata=float(fid.readline().strip().split()[1])
        if fmt==0: S.elev=loadtxt(fname,skiprows=6).astype('float32')
        fid.close()

        #shift half a cell if ll defined at corner
        #if xn.lower()=='xllcenter' and yn.lower()=='yllcenter': xll=xll+dxy/2; yll=yll+dxy/2;
        if xn.lower()=='xllcorner' and yn.lower()=='yllcorner': xll=xll+dxy/2; yll=yll+dxy/2;
        S.lon=xll+dxy*arange(ncols); S.lat=yll-dxy*arange(nrows)+(nrows-1)*dxy

    #save data
    if sname is not None: savez(sname,S)
    return S

def plot_taylor_diagram(R=None,STD=None,xm=2,cs=['b','k','g'],lss=['--','-','--'],lws=[0.5,0.5,1.0],
                        ticks=[None,None,None],fs=[8,8,8],lws_axis=[1.0,1.0],dxy=[0.005,0.01],stations=None,
                        labels=['Correlation Coefficient','Normalized Standard Deviation','RMSD'],fs_label=[12,12,12],R_label=[60,30,1.05]):
    '''
    plot taylor diagram, and return handles

    Input:
        R: correlation coefficient
        STD: normalized standard dievaiton (or standard dievation)
        xm: limit of x-axis and y-axis (this is axis for STD)
        cs/lss/lws/ticks/fs: list of colors/linestyles/lineweights/ticks/fontsizes for R, STD, and RMSD
        lws_axis: list of lineweights for xy-axis and outer rign (two values)

        labels: list of names for R,STD,RMSD
        fs_label: size of labels 
        R_label: parameters to set R labels (angle1, angle2, distance)
        dxy: set xlim=[-dxy[0],xm+dxy[0]],ylim=[-dxy[1],ym+dxy[1]]
        stations: when stations!=None, add legend

    note: after changing markers' properties, call self.hl.legend() to update legends
    '''

    #get default value for axis
    if ticks[0] is None: ticks[0]=array([*arange(0.1,1.0,0.1),0.95,0.99])
    if ticks[1] is None: ticks[1]=arange(0.5,5,0.5)
    if ticks[2] is None: ticks[2]=arange(0.5,5,0.5)
    sm=xm; S=zdata(); S.ht_label=[]

    #plot axis R
    xi=linspace(0,sm,200)
    S.hp_R=[plot(xi*i,xi*sqrt(1-i**2),color=cs[0],ls=lss[0],lw=lws[0]) for i in ticks[0]]
    S.ht_R=[text(1.01*sm*i,1.01*sm*sqrt(1-i**2), '{}'.format(float(int(i*100))/100),fontsize=fs[0],color=cs[0]) for i in ticks[0]]

    #add text for R
    rs=linspace(*R_label[:2],len(labels[0])); rm=sm*R_label[2] 
    for ri,ti in zip(rs,labels[0]): text(rm*cos(ri*pi/180),rm*sin(ri*pi/180),ti,fontsize=fs_label[0],fontweight='bold',color=cs[0],rotation=ri-90)
    #S.ht_label.append(text(0.97*sm*cos(45*pi/180),0.97*sm*sin(45*pi/180),labels[0],fontsize=fs_label[0],fontweight='bold',color=cs[0],rotation=-45))

    #plot STD
    ri=linspace(0,pi/2,200);
    S.hp_STD=[plot(cos(ri)*i,sin(ri)*i,color=cs[1],ls=lss[1],lw=lws[1]) for i in ticks[1] if i<sm]
    S.ht_label.append(text(-0.1*sm,0.25*sm,labels[1],fontsize=fs_label[1],fontweight='bold',color=cs[1],rotation=90))

    #plot RMSD
    ri=linspace(0,pi,200); xi=cos(ri); yi=sin(ri)
    S.hp_RMSD=[]
    for i in ticks[2]:
        #line
        xii=xi*i+1; yii=yi*i; fpn=(sqrt(xii**2+yii**2)<sm)*(xii>=0)
        if sum(fpn)==0: continue
        hl=plot(xii[fpn],yii[fpn], color=cs[2],ls=lss[2],lw=lws[2])

        #text
        xiii=abs(xii-(sm-0.85*i)/sm); sid=pindex(xiii==min(xiii))[0]
        text(1.02*xii[sid],1.02*yii[sid],'{}'.format(i),color=cs[2],fontsize=fs[2],rotation=15)

        S.hp_RMSD.append(hl)
    S.ht_label.append(text(0.08*sm,0.88*sm,labels[2],color=cs[2],fontsize=fs_label[2],fontweight='bold',rotation=25))

    #plot axis
    S.hp_axis1=plot(array([0,0,1])*sm,array([1,0,0])*sm,color='k',ls='-',lw=lws_axis[0])
    ri=linspace(0,pi/2,200); S.hp_axis2=plot(cos(ri)*sm,sin(ri)*sm,color='k',ls='-',lw=lws_axis[1]); 

    #plot pts
    if (R is not None) and (STD is not None):
        S.hp_obs=plot(1,0,'k.',ms=12,label='obs')
        S.hp=[];
        for i,ri in enumerate(R):
            xi=ri*STD[i]; yi=sqrt(1-ri**2)*STD[i]
            hp=plot(xi,yi,'r.',ms=10,label='{}'.format(i))
            S.hp.append(hp)

    #note
    setp(gca(),yticks=ticks[1],xticks=[])
    yticks(fontsize=fs[1]); gca().tick_params(length=0)
    gca().spines['right'].set_visible(False)
    gca().spines['top'].set_visible(False)
    gca().spines['left'].set_visible(False)
    gca().spines['bottom'].set_visible(False)
    setp(gca(),xlim=[-dxy[0],sm+dxy[0]],ylim=[-dxy[1],sm+dxy[1]])
    S.ha=gca(); S.ax=gca().axes;

    def update_legend(self=S,**args):
        self.hl=self.ha.legend(**args)
    S.update_legend=update_legend

    #add legend
    if stations is not None:
        S.hl=S.ha.legend(fontsize=8)
        if hasattr(stations,'__len__'): [S.hl.get_texts()[i+1].set_text(lstr) for i,lstr in enumerate(stations)]

    return S
def get_subplot_position2(margin=[0.1,0.1,0.1,0.1],dxy=[0.05,0.05],ds=[3,4],**args):
    '''
    return subplot position. Based on and calling get_subplot_position, but using the margin as input. 
    Inut:
        margin=[left,right,up,down]: distance from left, right, up, and bottom edge
        for other arguments, see get_subplot_position
    Sample function call:
        [ps,pc]=get_subplot_position2(margin=[0.05,0.05,0.1,0.1],dxy=[0.00,0.00],ds=[3,4],dc=[0.01,0.005])
        ps=reshape(ps,(12,4)) #to make 3D dimension array to 2 dimension
        for imon in arange(12)+1:
            axes(position=ps[imon-1])
    '''
    left,right,up,down=margin
    rown,coln=ds
    xspan=(1-left-right-(coln-1)*dxy[0])/coln #get x-span for each subplot
    yspan=(1-up-down-(rown-1)*dxy[1])/rown #get y-span for each subplot
    p0=[left,1-up-yspan,xspan,yspan] #get the upper left subplot position
    return get_subplot_position(p0,dxy,ds,**args)

def get_subplot_position(p0,dxy,ds,dc=None,sindc=None,figsize=None):
    '''
    return subplot position
    Input:
       p0=[x0,y0,xm,ym]: upper left subplot position
       dxy=[dx,dy]:      space between subplots
       ds=[ny,nx]:       subplot structure
       dc=[xmc,dxc]:     add colorbar position with width of xmc, and distance dxc from axes
       sindc=[:nplot]:   indices of subplot colorbars
       fsize=[fw,fh]:    plot out subplot in figure(figsize=fsize)
    '''

    #compute subplot position
    x0,y0,xm,ym=p0; dx,dy=dxy; ny,nx=ds
    ps=array([[[x0+i*(xm+dx),y0-k*(ym+dy),xm,ym] for i in arange(nx)] for k in arange(ny)])
    if dc!=None:
       xmc,dxc=dc; pc=zeros(ds).astype('O');  pc[:]=0
       if sindc!=None: pc.ravel()[setdiff1d(arange(prod(ds)),array(sindc))]=0
       for k in arange(ny):
           for i in arange(nx):
               if pc[k,i]!=None: pc[k,i]=[x0+xm+i*(xm+dx)+dxc,y0-k*(ym+dy),xmc,ym]

    #plot subplots
    if figsize!=None:
       figure(figsize=figsize)
       for i in arange(nx):
           for k in arange(ny):
               axes(position=ps[k,i]); xticks([]); yticks([])
               #setp(gca(),xticklabels=[],yticklabels=[])
               if dc!=None:
                  if pc[k,i]!=None: axes(position=pc[k,i]); xticks([]); yticks([])
    if dc!=None:
       return [ps,pc]
    else:
       return ps

def close_data_loop(xi):
    '''
    constructe data loop along the first dimension.
    if xi[0,...]!=xi[-1,...], then,add xi[0,...] in the end
    '''
    if array_equal(xi[0,...].ravel(),xi[-1,...].ravel()):
        vi=xi
    else:
        vi=r_[xi,xi[0,...][None,...]]
    return vi

def isnumber(x):
    '''
    check whether a string is numeric or not
    '''
    try:
       float(x); return True
    except:
      return False

def find_cs(xi,dx):
    '''
    analyze time series to find the locations where differences are larger than dx: (xi[i+1]-xi[i]>dx)
    return: 
      bind: locations of differences>dx
      sections: continous sections where all differences are smaller than dx 
      gaps:     gap sections where differences are larger than dx 
      slen,glen: lens for each section or gap 
      msection,mgap: maximum section/gap
    '''
    sind=pindex(diff(xi)>dx); sections=[]; gaps,glen,mgap=[],[],[]
    if len(sind)==0:
       sx=[xi[0],xi[-1]]; sections.append(sx)
    else:
       sx=[xi[0],xi[sind[0]]]; sections.append(sx)
       for m in arange(len(sind)-1):
           sx=[xi[sind[m]+1],xi[sind[m+1]]]; sections.append(sx) 
           gx=[xi[sind[m]],xi[sind[m]+1]]; gaps.append(gx)
       sx=[xi[sind[-1]+1],xi[-1]]; sections.append(sx)
       gx=[xi[sind[-1]],xi[sind[-1]+1]]; gaps.append(gx)
    sections=array(sections); gaps=array(gaps)
    slen=diff(sections,axis=1); msection=sections[pindex(slen,slen.max())[0]]
    if len(gaps)!=0: glen=diff(gaps,axis=1); mgap=gaps[pindex(glen,glen.max())[0]]
    S=zdata(); S.bind,S.sections,S.msection,S.slen,S.gaps,S.mgap,S.glen=sind,sections,msection,slen,gaps,mgap,glen
    return S

def datenum(*args,fmt=0):
    '''
    usage: datenum(*args,fmt=[0,1])
       datenum(2001,1,1,10,23,0)       or datenum([[2001,1,1],[2002,1,5]])
       datenum('2001-01-01, 10:23:00') or datenum(['2001-01-1','2002-01-05'])
       datenum(datetime.datetime(2000, 1, 1, 0, 0)) or datenum([datetime.datetime(2000,...), datetime.datetime(2000,...)])
       fmt=0: output num; fmt==1: output date
    '''
    import datetime

    #input only one argument, it should be a time string
    if len(args)==1: args=args[0]

    if isinstance(args,str):
       #single time string
       dnum=datestr2num(args)
       if fmt!=0: dnum=num2date(dnum)
    elif isinstance(args,datetime.datetime):
       dnum=mpl.dates.date2num(args) if fmt==0 else args
    elif hasattr(args,"__len__"):
       if isinstance(args[0],str)|hasattr(args[0],"__len__")|isinstance(args[0],datetime.datetime):
          #array of time string or time array
          dnum=array([datenum(i,fmt=fmt) for i in args])
       else:
          dargs=[*args]
          #if month >12: convert
          if dargs[1]>12:
             dargs[0]=dargs[0]+int(dargs[1]/12)
             dargs[1]=max([1,dargs[1]%12])

          #time array (e.g. [2002,1,1])
          dnum=datetime.datetime(*dargs)
          if fmt==0: dnum=mpl.dates.date2num(dnum)
    else:
       sys.exit('unknown input format')

    return dnum

def doy(daynum):
    '''
    return day of the year for datenum
    '''
    t0=array([datenum(num2date(i).year,1,1) for i in daynum]) if hasattr(daynum,'__len__') else datenum(num2date(daynum).year,1,1)
    return daynum-t0

def quickdatenum(times):
    '''
    To quickly process time stamps, suitable for data with continuously increasing time. It will save
    a lot of time when the number of record is large.
    Algorithm: If difference in timing between first and last records are equal to (n-1)*(t1-t0),
    where t0 and t1 is the first and second record,
    then arange(t0,tend,t1-t0) will be used as the time for the entire block.
    A recursive call is used when the above condition is not met.
    '''
    if len(times)==1: return(datenum(times))
    t0=datenum(times[0]); t1=datenum(times[1]); tend=datenum(times[-1])
    if abs(tend-t0-(len(times)-1)*(t1-t0))<(t1-t0)*1e-5 and t1!=t0:
        nums=arange(t0,tend+(t1-t0)*1e-5,t1-t0)
        if len(nums) != len(times):
            print(len(nums),len(times),times)
            raise ValueError('something is wrong')
        return nums
    else:
        midp=round(len(times)/2) #middle index to split the array
        #print(len(times),times) #for some diagonistic purpose to check where timing is not continuous
        return concatenate([quickdatenum(times[:midp]),quickdatenum(times[midp:])]) #recursive

def get_xtick(fmt=0,xts=None,str=None):
    '''
    see xtick's help doc
    '''
    return xtick(xts,fmt,str)

def xtick(xts=None,fmt=0,str=None):
    '''
    return temporal ticks and labels for plot purpose

        fmt: format of xtick and xticklabel
             0: year; 1: month; 2: day;  3: user-defined
        xts: time aranges or time arrays
            e.g. [2000,2010],[datenum(2000,1,1),datenum(2010,1,1)],[*arange(730120,730420)]
        str: format of label
            Year:   %y=01;  %-y=1;   %Y=2000
            Month:  %m=01;  %-m=1;   %B=January;  %b=Jan
            Day:    %d=01;  %-d=1
            Hour:   %H=09;  %-H=9;   %I=[00,12];  %-I=1
            Minute: %M=09;  %-M=9;
            Second: %S=09;  %-S=9
            AM/PM:  %p=[AM,PM]

            %c='Fri Jan 25 04:05:02 2008'
            %x='01/25/08'
            %X='04:05:02'

            Week:           %a=MON;      %A=Monday;   %w=[0,6]
            Week of year:   %U=[00,53];  %W=[00,53]
            Day of year:    %j=045;       %-j=45

            1: 2008-02-03
            2: 2008-02-03, 04:05:00
            3: J,F,M,A,J,J,... (Months)
            4： 03/15 (mm/dd)
    '''
    #get time label
    ft=0
    if str==1: str='%Y-%m-%d'
    if str==2: str='%Y-%m-%d, %H:%M:%S'
    if str==3: str='%b'; ft=1
    if str==4: str='%m/%d'

    #get time ticks
    it=0
    if xts is None:
        if fmt==3: sys.exit('must provide xts for fmt=3')
        if fmt==2:
            xts=[2000,2000]
        else:
            xts=[*arange(2000,2025)]
    elif len(xts)==2: #[year1,year2] or [datenum1, datenum2]
        if xts[0]>1e4: it=1
        xts=[*arange(xts[0],xts[1]+1)]
    else: #array of datenum
        it=1

    #compute time ticks and ticklabels
    if fmt==0:   #for year
        if str is None: str='%Y'
        if it==0: xts=[datenum(i,1,1) for i in xts]
    elif fmt==1: #for months
        if str is None: str='%b'
        if it==0: xts=array([[datenum(i,k+1,1) for k in arange(12)] for i in xts]).ravel()
    elif fmt==2: #for days
        if str is None: str='%-d'
        if it==0: xts=[*arange(datenum(min(xts),1,1),datenum(max(xts)+1,1,1)-1)]
    elif fmt==3: #user defined
        if str is None: str='%-d'
    xls=[num2date(i).strftime(str) for i in xts]
    if ft==1: xls=[i[0] for i in xls]

    return [xts,xls]

#-------loadz------------------------------------------------------------------
def get_INFO(data,fmt=0):
    '''
    collect information about object's attributes
    fmt=0: only return attributes; fmt=1: also return methods
    '''
    atts=[]; sdict=data.__dict__; skeys=sdict.keys(); fnc=0
    skeys=[i for i in skeys if not hasattr(sdict[i],'__call__')] #exclude method
    if ('file_format' in skeys) and ('netcdf' in sdict['file_format'].lower()) and fmt==0: #exclude default netcdf attribute
       skeys=[i for i in skeys if i not in ['name','groups','disk_format','path','parent','data_model','cmptypes','vltypes','enumtypes','keepweakref','auto_complex']]
    if ('dimname' in skeys) and ('dims' in skeys) and ('file_format' in skeys): fnc=1 #netcdf
    stypes=[int,int8,int16,int32,int64, float,float16,float32,float64]
    snames=['int','int8','int16','int32','int64','float','float16','float32','float64']
    for i in skeys:
        try:
            vi=sdict[i]; dt=type(vi); dta=''
            if (fnc==1) and (dt is zdata): vi=vi.val; dt=np.ndarray #netcdf file
            #get data information
            if dt is list:
               nd=': list({},)'.format(len(vi))
               if (fnc==1) and (i in ['dimname','dims']): nd=': list{}'.format(vi) #netcdf file
            elif dt is dict:
               nd=': dict({},)'.format(len(vi))
            elif dt is str:
               nd=': "{}", string'.format(vi[:30])
            elif dt is np.ndarray:
               nd=': array{}'.format(vi.shape); dta=str(vi.dtype)
               if (fnc==1) and (i in ['dimname','dims']): nd=': array({})'.format(vi) #netcdf file
               if vi.size==1 and (vi.dtype in stypes): nd=': {}, array({})'.format(squeeze(vi),1)
            elif dt in stypes:
               nd=': {}, {} '.format(vi,snames[stypes.index(dt)])
            elif 'narray' in str(dt).lower(): #netcdf4 variable
               nd=': narray{}, {}'.format(vi.shape,vi.dtype)
            elif 'npzfile' in str(dt).lower(): #npz variable
               if i.startswith('_') and i.endswith('_variables') and i[1]!='_': continue
               nd=': array{}, {}'.format(vi.shape,str(vi.dtype))
            else:
               nd=': {}'.format(type(vi))

            #output
            ms=min([6,max([len(k) for k in skeys])]); fs1='{:'+str(ms)+'s}{}'; fs2=fs1+', {}'
            fstr=fs2.format(i,nd,dta) if dta!='' else fs1.format(i,nd)
            atts.append(fstr.strip())
        except:
            pass

    #collect method
    if fmt==1:
       mths=['method: '+i for i in data.__dir__() if (i not in ['INFO','VINFO',*skeys]) and (not i.startswith('_'))]
       if len(mths)!=0: atts.extend(mths)
    return atts

class zdata:
    '''
    self-defined data structure by Zhengui Wang.  Attributes are used to store data
    '''
    def __init__(self):
        pass

    @property
    def INFO(self):
        return get_INFO(self)
    @property
    def VINFO(self):
        return get_INFO(self,1)

    def init(self,svar,value=None):
        '''
        initilize variables with value
        '''
        svar=[svar,] if isinstance(svar,str) else svar
        for i in svar: self.__dict__[i]=value.copy() if ['copy' in value.__dir__()] else value

    def getattr(self,svar=None,value=None,fmt=0,default=None):
        '''
        retrieve attribute or add attribute
        retrieve attrs:
            1). C.attr('a'): return C.a
            2). C.attr(['a','b']): return [C.a,C.b] if fmt==0 else array([C.a,C.b])
            3). C.attr(): return list of attribute names
        add attr:
            1). C.attr('a',x): set C.a=x
            2). C.attr(['a','b'],['x','y']): set C.a=x and C.b=y
        '''
        sdict=self.__dict__; T=isinstance(svar,str); svars=[svar,] if T else svar
        if (svar is None):  return array([*sdict])
        if (value is None):
           vs=[sdict[i] if (i in sdict) else default for i in svars]; return vs[0] if T else vs if fmt==0 else array(vs)
        else:
           for i,k in enumerate(svars): sdict[k]=value if T else value[i]

    def delattr(self,*args):
        '''
        delattr('x'), delattr('x','y'), or delattr(['x','y'])
        '''
        for i in args: svars=[i,] if isinstance(i,str) else i; [delattr(self, m) for m in svars if self.hasattr(m)]

    def attr(self,*args,fmt=0,default=None):
        '''
        alias to getattr: retrieve attribute or add attribute
        '''
        return self.getattr(*args,fmt=fmt,default=default)

    def hasattr(self,*args,fmt=0):
        '''
        check whether variable(s) exists
        fmt=0: return True or False; fmt=1: return a list of True or False
        '''
        fs=[hasattr(self,i) if isinstance(i,str) else self.hasattr(*i,fmt=fmt) for i in args]
        return sum(fs)==len(fs) if fmt==0 else fs

    def to_array(self,*args,dtype=None):
        '''
        convert attributes to numpy array
        1).  to_array('x') or to_array('x',dtype='int32');   2). to_array(['x','y']) or to_array('x','y')
        '''
        if len(args)==0: args=self.attr()
        sdict=self.__dict__
        for arg in args: 
            svars=[arg,] if isinstance(arg,str) else arg
            for i in svars: sdict[i]=array(sdict[i]) if (dtype is None) else array(sdict[i],dtype)

    def rename(self,svar,tvar):
        '''
        rename variable name
        '''
        sdict=self.__dict__; sdict[tvar]=sdict.pop(svar)

    def save(self,fname,**args):
        '''
        save zdata in differnt format based on filename extension (*.npz,*.nc,*.shp)
        '''
        if fname.endswith('.shp'):
           write_shapefile_data(fname,self,**args)
        elif fname.endswith('.nc'):
           WriteNC(fname,self,fmt=1 if isinstance(self,ncfile) else 0, **args)
        else:
           savez(fname,self)

def savez(fname,data,fmt=0):
    '''
    save data as self-defined python format
       fmt=0: save data as *.npz (small filesize, reads slower)
       fmt=1: save data as *.pkl (large filesize, reads faster)
    if fname endswith *.npz or *.pkl, then fmt is reset to match fname
    '''

    #determine format
    if fname.endswith('.npz'): fmt=0; fname=fname[:-4]
    if fname.endswith('.pkl'): fmt=1; fname=fname[:-4]
    if fname.endswith('.pp'):  fmt=2
    if fmt==1: fname=fname+'.pkl'
    if isinstance(data,ncfile):
       dvs=['name','groups','dimensions','variables','disk_format','path','parent','file_format','data_model',
            'cmptypes','vltypes','enumtypes','keepweakref','auto_complex','dimname','dims','dim_unlimited']
       svars0=[i for i in data.__dict__ if (not hasattr(data.attr(i),'__call__')) and (i not in dvs)]
    elif type(data)!=zdata:
         import cloudpickle; data._CLASS=cloudpickle.dumps(type(data))

    #save data
    if fmt in [0,2]:
       #check variable types (list, string, int, float and method), and constrcut save_string
       zdict=data.__dict__;  svars=svars0 if isinstance(data,ncfile) else list(zdict.keys())
       save_str='savez_compressed("{}" '.format(fname)
       lvars=[]; tvars=[]; ivars=[]; fvars=[]; mvars=[]
       for svar in svars:
           vi=zdict[svar]
           if isinstance(vi,int):   ivars.append(svar)
           if isinstance(vi,float): fvars.append(svar)
           if isinstance(vi,str):   tvars.append(svar)
           if isinstance(vi,list): #change list to array
                save_str=save_str+',{}=array(data.{},dtype="O")'.format(svar,svar); lvars.append(svar)
           elif callable(vi): #change function to bytes
                try:
                   import cloudpickle
                   save_str=save_str+',{}=cloudpickle.dumps(data.{})'.format(svar,svar); mvars.append(svar)
                except:
                   pass
           else:
                save_str=save_str+',{}=data.{}{}'.format(svar,svar,'.value' if ('narray' in str(type(vi))) else '')
       exec(save_str+',_list_variables=lvars,_str_variables=tvars,_int_variables=ivars,_float_variables=fvars,_method_variables=mvars)')
       if fmt==2: os.rename(fname+'.npz',fname)
    elif fmt==1:
       import pickle
       fid=open(fname,'wb'); pickle.dump(data,fid,pickle.HIGHEST_PROTOCOL); fid.close()

class npzfile(zdata):
    '''
    wrapper for numpy *.npz file
    '''
    def __init__(self,fname):
        #collect all variable information
        import zipfile; fid=zipfile.ZipFile(fname); ff=np.lib.format; svars0,svars,dms,dts=[],[],[],[]
        [svars0.extend(ff.read_array(fid.open(i))) for i in fid.namelist() if i.endswith('_variables.npy')]
        for name in [i for i in fid.namelist() if i.endswith('npy') and (i[:-4] not in svars0) and not i.endswith('_variables.npy')]:
            npy=fid.open(name); dm,_,dt=ff._read_array_header(npy, ff.read_magic(npy))
            svars.append(name[:-4]); dms.append(dm); dts.append(dt)
        fid.close(); [self.attr(i,read(fname,i)) for i in svars0] #read some variables

        #assign each numpy array
        data=load(fname,allow_pickle=True,mmap_mode='r')
        for svar,dm, dt in zip(svars,dms,dts): self.attr(svar,read(fname,svar) if str(dt)=='object' else ntype(data,0,name=svar,vshape=dm,vtype=dt))

def loadz(fname,svars=None):
    '''
    load self-defined data "fname.npz" or "fname.pkl"
       1). svars=list of variables: only read certain variables; 2). svars='vars': return avaiable variables
       3). svars=1: open channel to npz file (read data later).
    '''
    if isinstance(fname,str) and ('.' not in fname): fname=fname+'.npz' #default file format
    if isinstance(fname,np.lib.npyio.NpzFile) or fname.endswith('.npz') or fname.endswith('.pp'):
       if svars==1: return npzfile(fname)
       #get data info, collect variables
       data0=fname if isinstance(fname,np.lib.npyio.NpzFile) else load(fname,allow_pickle=True)
       fmt=1 if isinstance(svars,str) else 0 #determine what to return
       if svars=='vars': return array([i for i in data0 if not i.endswith('_variables')]) #return variables list
       svars0=list(data0.keys()); svars=svars0 if svars is None else [svars] if isinstance(svars,str) else svars
       vlist=['_int_variables','_float_variables','_str_variables','_list_variables','_method_variables','_CLASS']
       ivars=list(data0[vlist[0]]) if (vlist[0] in svars0) else []
       fvars=list(data0[vlist[1]]) if (vlist[1] in svars0) else []
       tvars=list(data0[vlist[2]]) if (vlist[2] in svars0) else []
       lvars=list(data0[vlist[3]]) if (vlist[3] in svars0) else []
       mvars=list(data0[vlist[4]]) if (vlist[4] in svars0) else []

       #extract data
       vdata=zdata()
       if '_CLASS' in svars:
          try:
              import pickle; vdata=pickle.loads(data0['_CLASS'])()
          except:
              pass
       vdict=vdata.__dict__
       for svar in setdiff1d(svars,vlist):
           vi=data0[svar] #get variable
           if vi.dtype==dtype('O'): vi=vi[()]    #restore object
           if svar in ivars: vi=int(vi)          #restore int variable
           if svar in fvars: vi=float(vi)        #restore int variable
           if svar in tvars: vi=str(vi)          #restore str variable
           if svar in lvars: vi=vi.tolist()      #restore list variable
           if svar in mvars: #restore function
              try:
                 import pickle; vi=pickle.loads(vi)
              except:
                 continue
           vdict[svar]=vi
    elif fname.endswith('.pkl'):
       import pickle
       from copy import deepcopy as dcopy
       vdata=zdata(); fid=open(fname,'rb'); fmt=0
       data=pickle.load(fid)
       vdata.__dict__=dcopy(data).__dict__.copy()
       fid.close()
    else:
       sys.exit('unknown format: {}'.format(fname))
    return vdata if fmt==0 else vdata.__dict__[svars[0]]

def least_square_fit(X,Y,order=None):
    '''
    perform least square fit
    usage: CC,fit_data=least_square_fit(X,Y)
        X: data maxtrix (npt,n)
        Y: data to be fitted (npt)
        order: if X is 1D array, create polynomial maxtrix [npt,order]
    where CC(n) is coefficient, fit_data is data fitted
    '''
    if X.shape[0]!=len(Y): X=X.T
    if X.ndim==1 and (order is not None):
       X=array([X**i for i in arange(order)]).T
       print(X.shape)
    CC=inv(X.T@X)@(X.T@Y); fy=X@CC

    return [CC,fy]

#-------mfft-------------------------------------------------------------------
def mfft(xi,dt):
    '''
    Perform FFT for a time series, with a time interval specified

    usage: period,afx,pfx=mfft(xi,dt)
    input:
       xi: time series
       dt: time interval

    output:
       period[period],afx[amplitude],pfx[phase]
    '''
    from scipy.fftpack import fft, ifft

    N=xi.size;
    fx=fft(xi);
    afx=abs(fx[1:N//2])*2.0/N;
    pfx=angle(fx[1:N//2]);
    period=dt*N/arange(1,N//2);
    return period,afx,pfx

def command_outputs(code,shell=True):
    '''
    Capture the command output from the system

    usage:
         S=command_outputs('ls')
         print(S.stderr)
         print(S.stdout) # normally this is the results
    '''
    import subprocess
    p=subprocess.Popen(code,stdout=subprocess.PIPE,stderr=subprocess.STDOUT,shell=shell)
    stdout,stderr=p.communicate()
    if stdout!=None: stdout=stdout.decode('utf-8')
    if stderr!=None: stderr=stderr.decode('utf-8')

    out=zdata()
    out.stdout=stdout
    out.stderr=stderr
    return out

def sub_lines(xy,fmt=0):
    '''
    fmt=0: decompose a line seperated by nan to sections
    fmt=1/2: decompose a line/polygon seperated to sides (x1,y1,x2,y2)
    '''
    if fmt==0:
       if isnan(xy[0,0]): xy=xy[1:] #remove 1st nan
       if not isnan(xy[-1,0]): xy=r_[xy,nan*zeros([1,2])] #add nan in the end
       inn=pindex(isnan(xy[:,0])); nn=len(inn); xys=zeros(nn,'O'); xys[0]=xy[:inn[0]]
       for i in arange(nn-1): xys[i+1]=xy[(inn[i]+1):inn[i+1]]
       return xys
    elif fmt in [1,2]:
       xys=sub_lines(xy) if fmt==1 else sub_polygons(xy); sides=[]
       for xy in xys: sides.extend(c_[xy[:-1],xy[1:]])
       return array(sides)

def sub_polygons(xy):
    '''
    decompose a polygon (xy) seperated by nan to small polygons
    '''
    xys=sub_lines(xy)
    for i, xy in enumerate(xys): xys[i]=close_data_loop(xy)
    return xys

def near_pts(pts,pts0,method=0,N=100):
    '''
    return index of pts0 that pts is nearest
       usage: sind=near_pts(pts,pts0)
       pts0: c_[x0,y0];  pts: c_[x,y]  (or, pts=x+1j*y)
       algorithm: using sp.spatial.cKDTree (default)

    old methods: method=1 and  method=2
       usage: sind=near_pts(pts,pts0,method=1[2],N=100)
       pts[n,2]: xy of points
       pts0[n,2]: xy of points

       method=1: quick method by subgroups (N);
       method=2: slower methods
    '''

    if method==0:
        pts=c_[pts.real,pts.imag] if pts.ndim==1 else pts; pts0=c_[pts0.real,pts0.imag] if pts0.ndim==1 else pts0
        sind=sp.spatial.cKDTree(pts0).query(pts)[1]
    elif method==1:
        p=pts[:,0]+(1j)*pts[:,1]
        p0=pts0[:,0]+(1j)*pts0[:,1]

        # N=min(N,len(p)-1);
        # #--divide pts into subgroups based on dist, each group size is about N---------
        # ps0=[]; ps=[]; ds=[]; inds=[]
        # i0=0; mval=1e50; pflag=p==None
        # while(True):
        #     ps0.append(p[i0]);
        #     dist=abs(p-p[i0]); dist[pflag]=mval;
        #     ind=argsort(dist); sdist=dist[ind]; mds=sdist[N]; i0=ind[N]
        #     fp=dist<mds; psi=p[fp]; dsi=max(dist[fp]); pflag[fp]=True;
        #     ps.append(psi); ds.append(dsi); inds.append(pindex(fp))
        #     if mds==mval: break

        N=min([N,len(p)]);
        #--divide pts into subgroups based on dist, each group size is about N---------
        ps0=[]; ps=[]; ds=[]; inds=[]; inum=arange(len(p))
        while(True):
            if len(inum)==0: break
            dist=abs(p[inum]-p[inum[0]]); sind=argsort(dist); inum=inum[sind]; dist=dist[sind]; sN=min([N,len(inum)])
            ps0.append(p[inum[0]]); ps.append(p[inum[:sN]]); ds.append(dist[sN-1]); inds.append(inum[:sN])
            inum=inum[sN:]

        #---find nearest pts for each subgroup-----------------------------------------
        inds0=[]
        for m in arange(len(ps)):
            dist=abs(p0-ps0[m]); dsi=ds[m]; psi=ps[m]
            #---find radius around ps0[m]
            dsm=dsi
            while(True):
                fp=dist<=dsm;
                if sum(fp)>0:
                    break;
                else:
                    dsm=dsm+dsi;
            #-----subgroup pts of p0---------------------------------
            fp=dist<=(dsm+2*dsi); ind0=pindex(fp); p0i=p0[ind0];
            psii=psi[:,None]; p0ii=p0i[None,:]
            dist=abs(psii-p0ii); indi=dist.argmin(axis=1);
            inds0.append(ind0[indi]);

        #-arrange index-----------------
        pind=array([]).astype('int'); pind0=array([]).astype('int')
        for m in arange(len(inds)):
            pind=r_[pind,inds[m]]
            pind0=r_[pind0,inds0[m]]

        ind=argsort(pind);
        sind=pind0[ind]
    elif method==2:
        n=pts.shape[0]; n0=pts0.shape[0]
        N=max([min([1e7//n0,n]),1e2])
        print('total pts: {}'.format(n));

        i0=int(0); i1=int(N);
        while(True):
            print('processing pts: {}-{}'.format(i0,i1));
            x=pts[i0:i1,0]; y=pts[i0:i1,1]
            x0=pts0[:,0]; y0=pts0[:,1]
            dist=(x[None,:]-x0[:,None])**2+(y[None,:]-y0[:,None])**2

            indi=[];
            for i in arange(x.shape[0]):
                disti=dist[:,i];
                indi.append(pindex(disti,min(disti))[0])

            if i0==0:
                ind=array(indi);
            else:
                ind=r_[ind,squeeze(array(indi))];
            #next step
            i0=int(i0+N); i1=int(min([i1+N,n]))
            if i0>=n: break
        sind=ind

    return sind

def inside(xy,pxy,fmt=0):
    '''
    fmt=0: return indices of points that are inside polygons; fmt=1: return boolean value for each pts
    xy:  coordinates of inquiry pts (npt,2); nan is allowed
    pxy: coordinates of polygons (npt,2); mulitiple polygons seperated by nan are allowed 
    '''
    sindp=pindex(mpl.path.Path(pxy).contains_points(xy))
    if fmt==0:
       return sindp
    else:
       fp=zeros(len(xy),'bool'); fp[sindp]=True; return fp

def inside_polygon(pts,px,py,fmt=0,method=0):
    '''
    check whether points are inside polygons

    usage: sind=inside_polygon(pts,px,py):
       pts[npt,2]: xy of points
       px[npt] or px[npt,nploy]: x coordiations of polygons
       py[npt] or py[npt,nploy]: y coordiations of polygons
       (npt is number of points, nploy is number of polygons)

       fmt=0: return flags "index[npt,nploy]" for whether points are inside polygons (1 means Yes, 0 means No)
       fmt=1: only return the indices "index[npt]" of polygons that pts resides in
             (if a point is inside multiple polygons, only one indice is returned; -1 mean pt outside of all Polygons)

       method=0: use mpl.path.Path
       method=1: use ray method explicitly

    note: For method=1, the computation time is proportional to npt**2 of the polygons. If the geometry
          of polygons are too complex, dividing them to subregions will increase efficiency.
    '''
    #----use ray method-----------------------------
    #check dimension
    if px.ndim==1:
       px=px[:,None]; py=py[:,None]

    #get dimensions
    npt=pts.shape[0]; nv,npy=px.shape

    if nv==3 and fmt==1:
       #for triangles, and only return indices of polygons that points resides in
       px1=px.min(axis=0); px2=px.max(axis=0); py1=py.min(axis=0); py2=py.max(axis=0)

       sind=[];
       for i in arange(npt):
           pxi=pts[i,0]; pyi=pts[i,1]
           sindp=pindex((pxi>=px1)*(pxi<=px2)*(pyi>=py1)*(pyi<=py2)); npy=len(sindp)
           if npy==0:
               sind.append(-1)
           else:
               isum=ones(npy)
               for m in arange(nv):
                   xi=c_[ones(npy)*pxi,px[m,sindp],px[mod(m+1,nv),sindp]]
                   yi=c_[ones(npy)*pyi,py[m,sindp],py[mod(m+1,nv),sindp]]
                   area=signa(xi,yi)
                   fp=area<0; isum[fp]=0;
               sindi=pindex(isum!=0)

               if len(sindi)==0:
                   sind.append(-1)
               else:
                   sind.append(sindp[sindi[0]])
       sind=array(sind)
    else:
        if method==0:
            sind=[]
            for m in arange(npy):
                sindi=mpl.path.Path(c_[px[:,m],py[:,m]]).contains_points(pts)
                sind.append(sindi)
            sind=array(sind).T+0  #convert logical to int
        elif method==1  :
            #using ray method explicitly
            sind=ones([npt,npy])
            x1=pts[:,0][:,None]; y1=pts[:,1][:,None]
            for m in arange(nv):
                x2=px[m,:][None,:]; y2=py[m,:][None,:]; isum=zeros([npt,npy])
                # sign_x1_x2=sign(x1-x2)
                for n in arange(1,nv-1):
                    x3=px[(n+m)%nv,:][None,:]; y3=py[(n+m)%nv,:][None,:]
                    x4=px[(n+m+1)%nv,:][None,:]; y4=py[(n+m+1)%nv,:][None,:]

                    #formulation for a ray to intersect with a line
                    fp1=((y1-y3)*(x2-x1)+(x3-x1)*(y2-y1))*((y1-y4)*(x2-x1)+(x4-x1)*(y2-y1))<=0 #intersection inside line P3P4
                    fp2=((y2-y1)*(x4-x3)-(y4-y3)*(x2-x1))*((y4-y3)*(x1-x3)+(y3-y1)*(x4-x3))<=0 #P1, P2 are in the same side of P3P4

                    fp12=fp1*fp2
                    isum[fp12]=isum[fp12]+1
                fp=((isum%2)==0)|((x1==x2)*(y1==y2))
                sind[fp]=0

        #change format
        if fmt==1:
            sindm=argmax(sind,axis=1)
            sindm[sind[arange(npt),sindm]==0]=-1
            sind=sindm
        elif fmt==0 and npy==1:
            sind=sind[:,0]
    return sind

def mdist(xy1,xy2,fmt=0,outfmt=0):
    '''
      compute distance or (minimum distance) between geometry (pts,lines) to geomtry(pts,lines,polygons)
      format of xy:
        points: c_[x,y]
        lines:  c_[x,y], or c_[x1,y1,x2,y2];  multiple lines can be seperated by nan
        polygons: c_[x,y]; multiple polygons can be seperated by nan
      fmt: choose geometry types
          0: pts and pts;       1: pts and lines;        2: lines and lines
          3: pts and polygons;  4: lines and polygons;   5: polygons and polygons
      outfmt=0: return an array of minimum distance;  outfmt=1: return an matrix of distance
    '''

    def pts_pts(xy,xy0,outfmt=0):
        '''
        find the minimum distance of c_[x,y] to a set of points c_[x0,y0]
          outfmt=0: minimum distance of pts to all pts (return an array)
          outfmt=1: minimum distance of pts to each pts (return a matrix)
        '''
        if outfmt==0:
            pid=near_pts(xy,xy0); dist=abs((xy[:,0]+1j*xy[:,1])-(xy0[pid,0]+1j*xy0[pid,1]))
        else:
            dist=abs((xy[:,0]+1j*xy[:,1])[:,None]-(xy0[:,0]+1j*xy0[:,1])[None,:])
        return dist

    def pts_lines(xy,lxy,outfmt=0):
        '''
        find the minimum distance of c_[x,y] to a set of lines lxy
          1). lxy=c_[x1,y1,x2,y2]: each row is a line (x1,y1) -> (x2,y2)
          2). lxy=c_[px,py]: lines are obtained with (px[i],px[i]) -> (px[i+1],px[i+1])
          outfmt=0: minimum distance of pts to all lines (return an array)
          outfmt=1: minimum distance of pts to each lines (return a matrix)
        '''

        #reshape pts and lines
        x,y=xy.T[:,:,None]; x1,y1,x2,y2=lxy.T[:,None,:]

        #compute the foot of a perpendicular, and distance between pts
        dist=nan*ones([x.size,x1.size])
        k=-((x1-x)*(x2-x1)+(y1-y)*(y2-y1))/((x1-x2)**2+(y1-y2)**2); xn=k*(x2-x1)+x1; yn=k*(y2-y1)+y1
        fpn=(k>=0)*(k<=1); dist[fpn]=abs((x+1j*y)-(xn+1j*yn))[fpn] #normal line
        dist[~fpn]=array(r_[abs((x+1j*y)-(x1+1j*y1))[None,...], abs((x+1j*y)-(x2+1j*y2))[None,...]]).min(axis=0)[~fpn] #pt-pt dist
        return dist.min(axis=1) if outfmt==0 else dist

    def lines_lines(lxy,lxy0,outfmt=0):
        '''
        find the minimum distance of lines lxy to a set of lines lxy0
          1). lxy or lxy0=c_[x1,y1,x2,y2]: each row is a line (x1,y1) -> (x2,y2)
          2). lxy or lxy0=c_[px,py]: lines are obtained with (px[i],px[i]) -> (px[i+1],px[i+1])
        outfmt=0: minimum distance of lines lxy to all lines lxy0 (return an array)
        outfmt=1: minimum distance of lines lxy to each lines in lxy0 (return a matrix)
        Note: if lines are intersecting, the minimum distance is zero.
        '''
        #reshape lines
        x1,y1,x2,y2=lxy.T[:,:,None]; x3,y3,x4,y4=lxy0.T[:,None,:]

        #check whether intersects
        fc1=(((x1-x3)*(y4-y3)+(x3-x4)*(y1-y3))*((x2-x3)*(y4-y3)+(x3-x4)*(y2-y3)))<=0
        fc2=(((x3-x1)*(y2-y1)+(x1-x2)*(y3-y1))*((x4-x1)*(y2-y1)+(x1-x2)*(y4-y1)))<=0
        fpn=(fc1*fc2); x1,y1,x2,y2=x1[:,0],y1[:,0],x2[:,0],y2[:,0]; x3,y3,x4,y4=x3[0],y3[0],x4[0],y4[0]

        if outfmt==1:
            dist=nan*ones([x1.size,x3.size]); dist[fpn]=0

            if sum(~fpn)!=0:
                #check distance between pts to lines
                dist[~fpn]=array([pts_lines(c_[x1,y1],c_[x3,y3,x4,y4],1),pts_lines(c_[x2,y2],c_[x3,y3,x4,y4],1),
                   pts_lines(c_[x3,y3],c_[x1,y1,x2,y2],1).T,pts_lines(c_[x4,y4],c_[x1,y1,x2,y2],1).T]).min(axis=0)[~fpn]
        else:
            dist=nan*ones(x1.size); fpn=fpn.sum(axis=1)>0; dist[fpn]=0
            x1,y1,x2,y2=x1[~fpn],y1[~fpn],x2[~fpn],y2[~fpn]
            if sum(~fpn)!=0:
                dist[~fpn]=array([pts_lines(c_[x1,y1],c_[x3,y3,x4,y4]),pts_lines(c_[x2,y2],c_[x3,y3,x4,y4]),
                   pts_lines(c_[x3,y3],c_[x1,y1,x2,y2],1).min(axis=0),pts_lines(c_[x4,y4],c_[x1,y1,x2,y2],1).min(axis=0)]).min(axis=0)
        return dist

    def pts_polygons(xy,pxy,outfmt=0):
        '''
        find the minimum distance of c_[x,y] to a polygon pxy
        Note: If pts are inside the polygon, the distance (minimum distance) is zero
        '''
        npt,npy=len(xy),len(pxy); dist=nan*zeros([npt,npy])

        #check pts in polygon
        for n, cxy in enumerate(pxy):
            fpn=inside(xy,cxy,1); dist[fpn,n]=0
            if sum(~fpn)!=0: dist[~fpn,n]=pts_lines(xy[~fpn],sub_lines(cxy,1))
        return dist.min(axis=1) if outfmt==0 else dist

    def lines_polygons(lxy,pxy,outfmt=0):
        '''
        find the minimum distance of lines lxy to a polygon pxy
        Note: if pts of lines are inside the polygon, the minimum distance is zero.
        '''
        npt,npy=len(lxy),len(pxy); x1,y1,x2,y2=lxy.T; dist=nan*ones([npt,npy])
        for n, xy in enumerate(pxy): #check each polygon
            fpn=inside(c_[x1,y1],xy,fmt=1)|inside(c_[x2,y2],xy,fmt=1); dist[fpn,n]=0 #check pts inside the polygon
            if sum(~fpn)!=0: dist[~fpn,n]=lines_lines(lxy[~fpn],sub_lines(xy,1))
        return dist.min(axis=1) if outfmt==0 else dist

    def polygons_polygons(pxy,pxy0,outfmt=0):
        '''
        find minimum distance between polygons
        '''
        npy,npy0=len(pxy),len(pxy0); dist=nan*ones([npy,npy0])
        for m,xy in enumerate(pxy):
            dist[m]=pts_polygons(xy,pxy0,1).min(axis=0); fpn=dist[m]!=0
            if sum(fpn)!=0: dist[m,fpn]=lines_polygons(sub_lines(xy,1),pxy0[fpn],1).min(axis=0)
        return dist.min(axis=1) if outfmt==0 else dist

    #change format of lines to (x1,y1,x2,y2), and change format of polygons
    if fmt in [2,4]:   xy1=xy1 if xy1.shape[1]==4 else sub_lines(xy1,1) #line
    if fmt in [1,2]:   xy2=xy2 if xy2.shape[1]==4 else sub_lines(xy2,1) #line
    if fmt in [5,]:    xy1=sub_polygons(xy1) #polygon
    if fmt in [3,4,5]: xy2=sub_polygons(xy2) #polygon

    #compute minimum distance
    if fmt==0: _mdist=pts_pts
    if fmt==1: _mdist=pts_lines
    if fmt==2: _mdist=lines_lines
    if fmt==3: _mdist=pts_polygons
    if fmt==4: _mdist=lines_polygons
    if fmt==5: _mdist=polygons_polygons
    return _mdist(xy1,xy2,outfmt)

def signa(x,y):
    '''
        compute signed area for triangles along the last dimension (x[...,0:3],y[...,0:3])
    '''
    if x.ndim==1:
        area=((x[0]-x[2])*(y[1]-y[2])-(x[1]-x[2])*(y[0]-y[2]))/2
    elif x.ndim==2:
        area=((x[...,0]-x[...,2])*(y[...,1]-y[...,2])-(x[...,1]-x[...,2])*(y[...,0]-y[...,2]))/2
    #area=np.squeeze(area)
    return area

def mdivide(A,B):
    '''
        perform matrix division B/A
    '''
    if A.ndim==1: A=A[None,:]
    if B.ndim==1: B=B[None,:]
    A2=inv(A@A.T);
    B2=B@A.T
    return squeeze(B2@A2)

def bpfilt(data,delta_t,band_f):
    '''
    band pass filter for 1D (data[time]) or nD (data[time,...]) array along the first dimension
       delta_t: time interval
       band_f: frequency band (eg. band_f=[12,48])
    '''
    from scipy.fftpack import fft, ifft

    #data dimension
    ds=data.shape; N=ds[0]

    #remove mean, do fft
    data=data-data.mean(axis=0)[None,...]
    fdata=fft(data,axis=0)

    #design filter
    filt=ones(N)
    k1=int(floor(band_f[0]*N*delta_t))-1; k1=max([min([k1,N//2]),1])
    k2=int(ceil(band_f[1]*N*delta_t))+1;  k2=max([min([k2,N//2]),1])
    filt[:k1]=0; filt[-k1:]=0; filt[k2:N-k2]=0
    filt=(ones([*ones(data.ndim).astype('int')])*filt).T

    #remove low and high freqs
    fdata=fdata*filt

    #bp results
    bpdata=real(ifft(fdata,axis=0))

    return bpdata

def lpfilt(data,delta_t,cutoff_f):
    '''
    low pass filter for 1D (data[time]) or nD (data[time,...]) array along the first dimension

    Note: there is no phase shift for this LP-filter
    '''
    #import gc #discard
    from scipy.fftpack import fft, ifft

    ds=data.shape

    #fft original data
    mdata=data.mean(axis=0)[None,...]
    #print(data.shape,mdata.shape)
    data=data-mdata
    fdata=fft(data,axis=0)

    #desgin filter
    N=ds[0];
    filt=ones(N)
    k=int(floor(cutoff_f*N*delta_t))
    filt[k]=0.715
    filt[k+1]=0.24
    filt[k+2]=0.024
    filt[k+3:N-(k+4)]=0.0
    filt[N-(k+4)]=0.024
    filt[N-(k+3)]=0.24
    filt[N-(k+2)]=0.715

    #expand dimension of filt
    filt=(ones([*ones(data.ndim).astype('int')])*filt).T

    #remove high freqs
    fdata=fdata*filt

    #lp results
    lfdata=real(ifft(fdata,axis=0))+mdata

    return lfdata

def smooth(xi,N):
    '''
    smooth average (on the 1st dimension):
       xi[time,...]: time series
       N: window size (if N is even, then N=N+1)
    '''

    #window span
    if mod(N,2)==0: N=N+1
    nz=int((N-1)/2)

    #moving averaging
    X=xi.copy(); SN=N*ones(xi.shape)
    for nmove in arange(1,nz+1):
        #sum in the beginning and end
        X[nmove:,...]=X[nmove:,...]+xi[:-nmove,...]
        X[:-nmove,...]=X[:-nmove,...]+xi[nmove:,...]

        #count
        SN[:nmove,...]=SN[:nmove,...]-1
        SN[-nmove:]=SN[-nmove:]-1
    SX=X/SN
    return SX

def daytime_length(lat,doy):
    '''
    calculate daytime length based on latitutde and day_of_year
    lat: latitude, doy: (1-365), sunrise=12-daytimelength/2, sunset=12+daytimelength/2
    '''
    P=arcsin(0.39795*cos(0.2163108 + 2*arctan(0.9671396*tan(0.00860*(doy-186)))));
    T=(sin(0.8333*pi/180)+sin(lat*pi/180)*sin(P))/cos(lat*pi/180)/cos(P);
    dt=24-(24/pi)*arccos(T);
    return dt

def move_figure(x=0,y=0,f=None):
    '''
    Move figure (f=gcf()) to upper left corner to pixel (x, y)
    e.g. move_figure(0,0,gcf())
    '''
    if f is None: f=gcf()
    backend = matplotlib.get_backend()
    if backend == 'TkAgg':
        f.canvas.manager.window.wm_geometry("+%d+%d" % (x, y))
    elif backend == 'WXAgg':
        f.canvas.manager.window.SetPosition((x, y))
    else:
        # This works for QT and GTK
        # You can also use window.setGeometry
        f.canvas.manager.window.move(x, y)

def modify_figure(fmt=0,hf=None):
    '''
    todo: not working yet !!
    alias for "gcf().tight_layout(); mvfig(); show(block=False)"
    '''
    if hf is None: hf=gcf()
    try:
       hf.tight_layout(); mvfig(f=hf)
       plt.ion(); show()
       #if fmt==0: hf.show(block=False)
       #if fmt==1: hf.show()
    except:
       pass

def fig_IFNO(hf=None,ax=None):
    '''
    get figure information
    '''
    if hf is None: hf=gcf()
    if ax is None: ax=gca()
    f='figure(figsize=[{},{}])'.format(hf.get_figwidth(),hf.get_figheight())
    a='xm={}; ym={}'.format(ax.get_xlim(),ax.get_ylim())
    b='setp(gca(),xticks=[],yticks=[],xlim=xm,ylim=ym)'
    c='gcf().tight_layout()'; d='show(block=False); mvfig()'
    print('\n'.join([f,a,b,c,d]))

def pause(dt=None):
    '''
    replace mpl.pylplot.pause for interaction
    '''
    if dt is None:
       try:
          gcf().canvas.draw()
       except:
          pass
       input()
    else:
       plt.pause(dt)

def proj(fname0=None,fmt0=None,prj0=None,fname1=None,fmt1=None,prj1=None,x=None,y=None,lon0=None,lat0=None,order0=0,order1=0):
    '''
    tranfrom projection of files: proj(fname0,fmt0,prj0,fname1,fmt1,prj1,x,y,lon0,lat0,order0,order1)
       fname: file name
       fmt: 0: SCHISM gr3 file; 1: SCHISM bp file; 2: xyz file; 3: xyz file with line number
       prj: projection name (e.g. 'epsg:26918', 'epsg:4326','cpp'), or projection string (e.g. prj0=get_prj_file('epsg:4326'))
       lon0,lat0: center for transformation between lon&lat and x&y; if lon0=None or lat0=None, then x0=mean(lon), and y0=mean(lat)
       order=0 for projected coordinate; order=1 for lat&lon (if prj='epsg:4326', order=1 is applied automatically')

    tranform data directly: px,py=proj(prj0='epsg:26918',prj1='epsg:4326',x=px,y=py)

    function used:
          lat,lon=Transformer.from_crs('epsg:26918','epsg:4326').transform(x,y)
          x,y=Transformer.from_crs('epsg:4326','epsg:26918').transform(lat,lon)
    #x1,y1=transform(Proj(proj0),Proj(proj1),x,y); #not used anymore
    '''
    from pyproj import Transformer
    from .schism_file import read_schism_hgrid,read_schism_bpfile,schism_bpfile

    #read file
    if fmt0==0:
        gd=read_schism_hgrid(fname0)
        np=gd.np; x=gd.x; y=gd.y; z=gd.dp
    elif fmt0==1:
        gd=read_schism_bpfile(fname0)
        np=gd.nsta; x=gd.x; y=gd.y; z=gd.z
    elif fmt0==2:
        gd=loadtxt(fname0)
        np=gd.shape[0]; x=gd[:,0]; y=gd[:,1]; z=gd[:,2]
    elif fmt0==3:
        gd=loadtxt(fname0)
        np=gd.shape[0]; x=gd[:,1]; y=gd[:,2]; z=gd[:,3]
    else:
        if x is None or y is None: sys.exit('unknown format of input files: {}, {}'.format(fname0,fname1))

    #order of data
    prj0=prj0.lower(); prj1=prj1.lower(); icpp=0
    if prj0=='epsg:4326': order0=1
    if prj1=='epsg:4326': order1=1
    if 'cpp' in [prj0,prj1]: icpp=1; rearth=6378206.4

    #transform coordinate
    if icpp==1 and prj0=='cpp':
        if prj1!='epsg:4326': sys.exit('projection wrong: prj0={}, prj1={}'.format(prj0,prj1))
        if (lon0 is None) or (lat0 is None): sys.exit('need lon0 and lat0 for cpp=>ll transform')
        x1=lon0+x*180/(pi*rearth*cos(lat0*pi/180)); y1=y*180/(pi*rearth)
    elif icpp==1 and prj1=='cpp':
        if prj0!='epsg:4326': sys.exit('projection wrong: prj0={}, prj1={}'.format(prj0,prj1))
        if lon0 is None: lon0=mean(x)
        if lat0 is None: lat0=mean(y)
        x1=rearth*(x-lon0)*(pi/180)*cos(lat0*pi/180); y1=rearth*y*pi/180
    else:
        if order0==1: x,y=y,x
        fpn=~(isnan(x)|isnan(y)); x1=arange(len(x))*nan; y1=arange(len(y))*nan
        x1[fpn],y1[fpn]=Transformer.from_crs(prj0,prj1).transform(x[fpn],y[fpn])
        if order1==1: x1,y1=y1,x1
        if (sum(isnan(x1[fpn]))!=0) | (sum(isnan(y1[fpn]))!=0): sys.exit('nan found in tranformation: x1,y1') #check nan

    # if (sum(isnan(x))!=0) and (sum(isnan(y))!=0): sys.exit('nan found in x,y') #check nan
    #x1,y1=Transformer.from_crs(prj0,prj1).transform(x,y)
    #x1,y1=transform(Proj(init=proj0),Proj(init=proj1),x,y);
    #x1,y1=transform(Proj(proj0),Proj(proj1),x,y);

    #write file
    if fmt1==0:
        if fmt0!=0: sys.exit('{} should have gr3 format'.format(fname0))
        gd.x=x1; gd.y=y1;
        gd.write_hgrid(fname1,Info='!coordinate transformed: {}=>{}'.format(prj0,prj1))
    elif fmt1==1:
        if fmt0==1:
            gd.x=x1; gd.y=y1
        else:
            gd=schism_bpfile()
            gd.note='coordinate transformed: {}=>{}'.format(prj0,prj1)
            gd.nsta=np; gd.x=x1; gd.y=y1; gd.z=z;
        gd.write_bpfile(fname1)
    elif fmt1==2 or fmt1==3:
        with open(fname1,'w+') as fid:
            for i in arange(np):
                if fmt1==2: fid.write('{} {} {}\n'.format(x1[i],y1[i],z[i]))
                if fmt1==3: fid.write('{} {} {} {}\n'.format(i+1,x1[i],y1[i],z[i]))
    else:
       return [x1,y1]

def proj_pts(x,y,prj1='epsg:4326',prj2='epsg:26918',**args):
    '''
    convert projection of points from prj1 to prj2
      x,y: coordinate of pts
      prj1: name of original projection
      prj2: name of target projection
      when 'cpp' projection exists, lon0, and lat0 arguments can be provided
    '''
    if len(array(x).shape)==0: x=array([x,]); y=array([y,])
    px,py=proj(prj0=prj1,prj1=prj2,x=x,y=y,**args)
    return [px,py]

def get_prj_file(prjname='epsg:4326',fmt=0,prj_dir=r'D:\Work\Database\projection\prj_files',fname=None):
    '''
    return projection name or entire database (dict)
        fmt=0: get one projection
        fmt=1: return dict of projection database
        fmt=-1: process *.prj files from prj_dir

    #-------online method-----------------
    #function to generate .prj file information using spatialreference.org
    #def getWKT_PRJ (epsg_code):
    #     import urllib
    #     # access projection information
    #     wkt = urllib.urlopen("http://spatialreference.org/ref/epsg/{0}/prettywkt/".format(epsg_code))
    #     # remove spaces between charachters
    #     remove_spaces = wkt.read().replace(" ","")
    #     # place all the text on one line
    #     output = remove_spaces.replace("\n", "")
    #     return output
    '''

    #get location of database
    bdir=os.path.dirname(__file__)

    if fmt==0:
        S=loadz('{}/../scripts/prj.npz'.format(bdir))
        return S.prj[prjname]
    elif fmt==1:
        S=loadz('{}/../scripts/prj.npz'.format(bdir))
        return S.prj
    elif fmt==-1:
         if not os.path.exists(prj_dir): return
         pnames=glob('{}/*.prj'.format(prj_dir)); prj={}
         for pname in pnames:
             line=open(pname,'r').readline().strip()
             pn=os.path.basename(pname)[:-4].replace('.',':').lower()
             prj[pn]=line
         if fname is not None: C=zdata(); C.prj=prj; C.save(fname)
         return prj
    else:
        sys.exit('unknow fmt')

#----------------convert_matfile_format----------------------------------------
#def convert_matfile_format(file):
#    '''
#    convert a matlab data to self-defined python format
#    file: input, a directory or a matfile
#    '''
#    # eg. file="C:\Users\Zhengui\Desktop\Observation2\DWR\SFBay_DWRData_SSI.mat"
#    # file="D:\OneDrive\Python\tem.mat"
#
#    fname=[]
#    if os.path.isdir(file):
#        sfile=os.listdir(file)
#        for sfilei in sfile:
#            ename=sfilei.rstrip().split('.')[-1]
#            if ename=='mat':
#                fname.append(file+os.sep+sfilei)
#    else:
#        fname=[file];
#
#    fid=open('log_matfile_convert.txt','w+')
#    #convert mat format from v7.3 to v7
#    cdir=os.getcwd();
#    os.chdir('D:\OneDrive\Python')
#    import matlab.engine
#    eng = matlab.engine.start_matlab()
#
#    for fn in fname:
#        print('converting matfile: '+fn)
#        dname=os.path.dirname(fn)
#        bname=os.path.basename(fn).split('.')[0]
#        fnv7=dname+os.sep+bname+'_v7'
#        fnz=dname+os.sep+bname
#        eflag=eng.convert_matfile_format(fn,fnv7,nargout=1)
#        if eflag!=0:
#            print('convert flag is %d: %s\n' % (eflag, fn));
#            fid.write('convert flag is %d: %s\n' % (eflag,fn))
#            continue
#        read_mat(fnz,fnv7)
#        os.remove(fnv7+'.mat')
#    fid.close()
#    os.chdir(cdir)
#

#convert between MATLAB matfile and zdata
def npz2mat(npz_data,fname):
    '''
      convert Python *.npz file/data to Matlab *mat file
      Exmaples
        1. npz2mat('test.npz','test.mat')
        2. npz2mat(C,'test.mat') #C=loadz('test.npz')
    '''
    from scipy import io

    if isinstance(npz_data,str): npz_data=loadz(npz_data)
    sp.io.savemat(fname,npz_data.__dict__)

def read_mat(matfile,fname=None):
    '''
    Convert Matlab *.mat file to Python *.npz file (alias: mat2npz)
    Inputs:
      matfile: name of matlab file ('name.mat' or 'name')
      fname: name of *.npz file to be saved  ('name.npz' or 'name')

    Examples: (mat2npz is alias to read_mat)
      1. mat2npz('A.mat','A.npz')
    '''
    from scipy import io

    #check name
    if matfile.endswith('.mat'): matfile=matfile[:-4]

    #read matfile and convert
    C=sp.io.loadmat(matfile+'.mat',simplify_cells=True); S=zdata()
    for x in C.keys():
       if x.startswith('__') or x=='INFO':continue

       #for list of strings
       if isinstance(C[x],np.ndarray):
          if C[x].size!=1 and ('<U' in str(C[x].dtype)):
             C[x]=array([i.strip() for i in C[x]])
       S.__dict__[x]=C[x]
    if fname is not None: savez(fname,S)
    return S

def sindex(A,B=None,fmt=0,**args):
    '''
    return index in differnt cases
      1). sindex(A):   sA,sind_sort,sind_inverse,ncount=unique(A,return_index=True,return_inverse=True,return_counts=True,**args) 
      2). sindex(A,B): C,sindA,sindB=intersect1d(A,B,return_indices=True) 
      3). sindex(A,B,1):  return the index of A's element in B. The index will be nan if A's element is not in B
                          if A is a subset of B; sindB=sindex(A,B,1) => A=B[sindB] 
                         
    '''
    if B is None: #single input
       if fmt==0: 
          sA,sind_sort,sind_inverse,ncount=unique(A,return_index=True,return_inverse=True,return_counts=True,**args)
          return sA,sind_sort,sind_inverse,ncount 
    else: #two arrays
       if fmt==0:
          C,sindA,sindB=intersect1d(A,B,return_indices=True); return C,sindA,sindB 
       elif fmt==1:
          sA,inverse_A=unique(A,return_inverse=True)
          C,sind_A,sind_B=intersect1d(sA,B,return_indices=True)
          sind=-ones(len(sA),'int'); sind[sind_A]=sind_B; return sind[inverse_A]

def pindex(data,value=None,inverse=0):
    '''
    return True indices of boolean array
        data: array of boolean values or numeric values (need to compare with value)
        value: comparison criteria that can be nan, numerics, or string
        inverse=1: return the False indices

        Examples:  1). pindex(fp);  2). pindex(x,nan); 3). pindex(x,3.0); 4). pindex(x,'>3.0'); 5). pindex(x,'a')
    '''
    if value is None:
       fp=data
    elif isinstance(value,str):
       fp=eval('data'+value) if (value[0] in ['>','<','=']) else data==value
    elif isnan(value):
       fp=isnan(data)
    else:
       fp=data==value
    return nonzero(fp if inverse==0 else ~fp)[0]

def nindex(data,value=None):
    return pindex(data,value,1)

def cindex(index,shape):
    '''
    convert array index: same as unravel_index and ravel_multi_index
      1) cindex(id, ds):  convert flat index (1D) to indices (nD)
      2) cindex(c_[id1,id2,...], ds):  convert flat indices (nD) to index (1D)
    '''
    index=array(index)
    if index.ndim!=1 and index.shape[1]==len(shape): index=index.T

    if index.ndim==1:
       cid=[*unravel_index(index,shape)]
    else:
       cid=ravel_multi_index(index,shape)
    return cid

def EOF(data,npc=8,scale=0,center=True,method=0,rotation=0,npc_r=None,**args):
    '''
    EOF analysis
        data(time,...): data with 1st dimension as time, other dimes for space
        npc: number of principal components (PCs) returned
        scale=0: normalized by eigenvalue; scale=1: normalized by the value of time series of each PC.
        method=0: using eofs package; method=1: using xeofs package
        rotation=1: return Varimax-rotated EOF analysis
        npc_r: optional number of rotation EOFs (fewer than npc)
        Note: the mean value is removed (center=True) in the analysis at default.
    Outputs: (PC,CC,VC,solver)
        PC: principal components
        CC: coefficients of each PC
        VC: variation of each PC
        solver: EOF analysis solver, and all results can be derived from it (e.g., solver.reconstructedField(8))
    '''
    if method==0 and rotation==0:
       from eofs.standard import Eof
       solver=Eof(data,center=center,**args)
       PC=solver.eofs(neofs=npc,eofscaling=2)
       CC=solver.pcs(npcs=npc,pcscaling=1).T
       VC=solver.varianceFraction(npc)
    else:
       import xeofs as xe; import xarray as xr
       ds=['d{}'.format(i) for i in arange(data.ndim)]; xs={i:arange(k) for i,k in zip(ds,data.shape)}
       solver=xe.single.EOF(npc).fit(xr.Dataset(data_vars={'data': xr.DataArray(data,dims=ds)},coords=xs),'d0')
       if rotation==1: solver=xe.single.EOFRotator(n_modes=min([npc,npc_r])).fit(solver)
       PC=solver.components().to_array().data[0]
       CC=solver.scores().data
       VC=solver.explained_variance_ratio().data

    #normalize
    for m,[pc,cc] in enumerate(zip(PC,CC)):
        s=1 if cc.mean()>=0 else -1 #whether to reverse the sign
        rat=sqrt(sum(pc**2)/pc.size) if scale==0 else sqrt(cc.size/sum(cc**2))
        PC[m]=s*pc/rat; CC[m]=s*cc*rat
    return PC, CC, VC, solver

def REOF(data,npc=8,scale=0,center=True,npc_r=None,**args):
    '''
    Varimax-rotated EOF analysis; see input in EOF
    '''
    return EOF(data,npc,scale,center,npc_r=npc_r,rotation=1,**args)

def get_stat(xi_model,xi_obs,fmt=0):
    '''
    compute statistics between two time series
    x1, x2 must have the same dimension
    x1: model; x2: obs
    fmt=1: compute pvalue using scipy.stats.pearsonr

    #import matlab.engine
    #eng=matlab.engine.start_matlab()
    '''

    #compute statistical numbers
    x1=array(xi_model); x2=array(xi_obs); npt=len(x1); S=zdata()
    if npt==0: S.R,S.ME,S.MAE,S.RMSD=nan,nan,nan,nan; return S
    dx=x1-x2; std1=std(x1); std2=std(x2); mx1=mean(x1); mx2=mean(x2)
    S.R=corrcoef(x1,x2)[0,1] if (npt>=3 and std2!=0 and std1!=0) else nan #R
    S.ME=mean(dx)
    S.MAE=mean(abs(dx))
    S.RMSD=sqrt((dx**2).mean())
    S.std=std(dx)
    S.ms=1-sum(dx**2)/sum((abs(x1-mx2)+abs(x2-mx2))**2)
    if fmt==1: a,S.pvalue=sp.stats.pearsonr(x1,x2)
    S.std1=std1; S.std2=std2; S.npt=npt
    S.taylor=array([sqrt(mean((x1-x1.mean())**2))/S.std2,sqrt(((x1-x1.mean()-x2+x2.mean())**2).mean())/S.std2,S.R]) if (npt>=3 and std2!=0) else [0,0,0]
    return S

def interp_vertical(data0,zcor0,zcor):
    '''
    interpolate data in the vertical dimension; works for multi-dimensional data.
    format:
      data0([nz0,...]: can be mutli-dimensional, but the 1st dim must be vertical
      zcor0[nz0,...]:  z-coordiantes of data; it can have fewer dimenions of data
      zcor[nz,...]:    target z-coordiantes.
    where nz0 and nz are number of vertical layers

    Note: zs and zcor can have different number of dimensions, can be different from data.
          However, their dimensions should be consistent with data if ndim>1.
    '''

    #collect dimension info
    v0=data0; z0=zcor0; z=zcor; nz0=len(z0); nz=len(z); ds=[nz,*v0.shape[1:]]; ndim=v0.ndim

    #check zcor order and dimension
    dz=array(z0[0]-z0[-1])
    if dz.min()*dz.max()<0: sys.exit('zcor should be either decreasing or increasing for all pts')
    if dz.max()>0: v0=v0[::-1]; z0=z0[::-1]

    #interpolate in the vertical
    v=zeros(ds)*nan; sfp=ones(ds)==1
    for i in arange(ndim-z0.ndim): z0=z0[...,None] #z0=expand_dims(z0,axis=-1) expand z dims
    for i in arange(ndim-z.ndim):  z=z[...,None]   #z=expand_dims(z,axis=-1)
    fp=z<=z0[0][None,...];  afp=sfp*fp; v[afp]=(v0[0][None,...]*sfp)[afp]  #extend bottom
    fp=z>=z0[-1][None,...]; afp=sfp*fp; v[afp]=(v0[-1][None,...]*sfp)[afp] #extend surface
    for k in arange(nz0-1):
        z1,z2=z0[k][None,...],z0[k+1][None,...]; dz=z2-z1; fpz=dz==0; dz[fpz]=1 #exclude z1=z2
        if sum(~fpz)==0: continue
        rat=(z-z1)/dz; fp=(rat>=0)*(rat<=1); fp=fp*(~fpz)
        if sum(fp)==0: continue
        afp=fp*sfp; v[afp]=((1-rat)*v0[k][None,...]+rat*v0[k+1][None,...])[afp]
    return v

def read_shapefile_data(fname):
    '''
    read shapefile (*.shp) and return its content

    note:  works for pts and polygon only, may not work for other geomerties (need update in these cases)
    '''

    import shapefile as shp
    with shp.Reader(fname) as C:
        #----read shapefile----------------
        S=zdata();
        S.nrec=C.numRecords
        S.type=C.shapeTypeName

        #----read pts----------------------------------------------------------
        #works for pts and polygon, may not work for other geomerty (need update in this case)
        S.xy=[];
        for i in arange(S.nrec):
            xyi=array(C.shape(i).points);
            parti=array(C.shape(i).parts,dtype='int');
            #insert nan for delimiter
            #to get original index: ind=pindex(xyi[:,0],nan)-arange(len(parti));
            S.xy.append(insert(xyi,parti,nan,axis=0))
        S.xy=squeeze(array(S.xy,dtype='O'))

        #---read attributes----------------------------------------------------
        S.attname=array([C.fields[i][0] for i in arange(1,len(C.fields))]);
        stype=array([type(C.record()[m]) for m in S.attname])
        svalue=array(C.records(),dtype='O');
        S.attvalue=array(zeros(len(S.attname))).astype('O')
        for m in arange(len(S.attname)):
            S.attvalue[m]=svalue[:,m].astype(stype[m])
        S.atttype=stype

        #read prj file if exist---
        bdir=os.path.dirname(os.path.abspath(fname));
        bname=os.path.basename(fname).split('.')[0]
        prjname='{}/{}.prj'.format(bdir,bname)
        if os.path.exists(prjname):
            with open(prjname,'r') as fid:
                S.prj=fid.readline().strip()

    return S

def write_shapefile_data(fname,data,prj='epsg:4326',float_len=18,float_decimal=8):
    '''
    write shapefiles
      fname: file name
      data: data in zdata() format
      prj: projection

    example: 
       S=zdata(); S.type=='POINT'; S.prj='epsg:4326'
       S.xy=c_[slon[:],slat[:]]
       S.attname=['station']; S.attvalue=station[:]
       write_shp('test.shp',S)

    note: only works for geometry: POINT, POLYLINE, POLYGON
    '''

    import shapefile as shp
    S=data
    if S.type=='POINT':
       nrec=len(S.xy)
    elif S.type=='POLYLINE' or S.type=='POLYGON':
        nrec=len(S.xy) if (S.xy.dtype==dtype('O') or S.xy.ndim==3) else 1
    else:
        sys.exit('unknown type')
    if hasattr(S,'nrec') and  nrec!=S.nrec: sys.exit('nrec inconsistent')

    #---write shapefile---------
    with shp.Writer(fname) as W:
        W.autoBalance=1;
        #define attributes
        if hasattr(S,'attname'):
            if not hasattr(S.attvalue,'dtype'): S.attvalue=array(S.attvalue)
            if S.attvalue.dtype==dtype('O'):
                stype=[type(S.attvalue[m][0]) for m in arange(len(S.attname))]
            else:
                if S.attvalue.ndim==1:
                    stype=[type(S.attvalue[0])]
                elif S.attvalue.ndim==2:
                    stype=[type(S.attvalue[m][0]) for m in arange(len(S.attname))]

            for m in arange(len(stype)):
                npint,npfloat,npstr=[np.int,np.float,np.str] if hasattr(np,'int') else [int,float,str]
                if stype[m] in [npint,np.int8,np.int16,np.int32,np.int64]:
                    W.field(S.attname[m],'N')
                elif stype[m] in [npfloat,np.float16,np.float32,np.float64]:
                    W.field(S.attname[m],'F',float_len,float_decimal)
                elif stype[m] in [npstr,np.str_,np.bytes_]:
                    W.field(S.attname[m],'C',100)
                else:
                    print('attribute type not included: {}'.format(stype[m]))
                    sys.exit()
        else:
            W.field('field','C')
            W.record()

        #put values
        for i in arange(nrec):
            if S.type=='POINT': #point, W.multipoint(S.xy) is multiple pts features
                W.point(*S.xy[i])
            elif S.type=='POLYLINE':
                vali=S.xy[i] if (S.xy.dtype==dtype('O') or S.xy.ndim==3) else S.xy
                W.line(delete_shapefile_nan(vali,0)) #reorganize the shape of vali, and write
            elif S.type=='POLYGON':
                vali=S.xy[i] if (S.xy.dtype==dtype('O') or S.xy.ndim==3) else S.xy
                W.poly([[[*m] for m in k] for k in delete_shapefile_nan(vali,1)]) #reorganize the shape of vali and add polygons

            #add attribute
            if hasattr(S,'attname'):
                if S.attvalue.dtype==dtype('O'):
                    atti=[S.attvalue[m][i] for m in arange(len(stype))]
                else:
                    if S.attvalue.ndim==1:
                        atti=[S.attvalue[i]]
                    elif S.attvalue.ndim==2:
                        atti=[S.attvalue[m][i] for m in arange(len(stype))]
                W.record(*atti)

        #----write projection------------
        bname=os.path.basename(fname).split('.')[0]
        bdir=os.path.dirname(os.path.abspath(fname))
        prjs=S.prj if hasattr(S,'prj') else prj
        if len(prjs)<20: prjs=get_prj_file(prjs)
        fid=open('{}/{}.prj'.format(bdir,bname),'w+'); fid.write(prjs); fid.close()

def delete_shapefile_nan(xi,iloop=0):
    '''
    delete nan (head and tail), and get ind for the rest
    iloop: prepare for multiple polygons
    '''
    if xi.ndim==1:
        i1=0; i2=xi.shape[0]
        if isnan(xi[0]): i1=1
        if isnan(xi[-1]): i2=i2-1
        yi=xi[i1:i2]; ind=pindex(yi,nan)
    elif xi.ndim==2:
        i1=0; i2=xi.shape[0]
        if isnan(xi[0,0]): i1=1
        if isnan(xi[-1,0]): i2=i2-1
        yi=xi[i1:i2]; ind=pindex(yi[:,0],nan)

    #------reorganize-----------
    if len(ind)==0:
        #close the geomety
        if iloop==1: yi=close_data_loop(yi)

        vi=[yi];
    else:
        vi=[];
        yii=yi[:ind[0]];
        if iloop==1: yii=close_data_loop(yii)
        vi.append(yii)
        for m in arange(len(ind)-1):
            i1=ind[m]+1; i2=ind[m+1];
            yii=yi[i1:i2];
            if iloop==1: yii=close_data_loop(yii);
            vi.append(yii)
        yii=yi[(ind[-1]+1):];
        if iloop==1: yii=close_data_loop(yii)
        vi.append(yii)

    return vi

#---------------------netcdf---------------------------------------------------
class ncfile(zdata):
      '''
      wraper for netcdf Datafile(fname)
      '''
      def __init__(self,fname,mode='r'):
          if isinstance(fname,str): import netCDF4; fname=netCDF4.Dataset(fname,mode=mode)
          for i in [i for i in fname.__dir__() if (not i.startswith('_')) and (not ' ' in i)]:
              exec('self.{}=fname.{}'.format(i,i)) #get method and attributes
          for i in self.variables: self.__dict__[i]=ntype(self.variables[i],fmt=1,name=i) #get variables
          self.dimname=[*self.dimensions]; self.dims=[self.dimensions[i].size for i in self.dimname]
          self.dim_unlimited=[self.dimensions[i].isunlimited() for i in self.dimname]

      #refine method
      def __getitem__(self,key):
         return self.__dict__[key]
      def __setattr__(self,key,value):
          if not hasattr(self,key): self.__dict__[key]=value; return
          cvar=self.__dict__[key]
          if hasattr(cvar,'_data'):
             cvar._value=value
          else:
             cvar=value
      def __delattr__(self,key):
          self.__dict__.pop(key); self.variables.pop(key)

def ReadNC(fname,fmt=0,mode='r',order=0):
    '''
    read netcdf files, and return its values and attributes
        fname: file name

        fmt=0: reorgnaized Dataset with format of zdata
        fmt=1: return netcdf.Dateset(fname)
        fmt=2: reorgnaized Dataset (*npz format), ignore attributes
        fmt='vars': return available variables

        mode='r+': change variable in situ. other choices are: 'r','w','a'

        order=1: only works for med=2; change dimension order
        order=0: variable dimension order read not changed for python format
        order=1: variable dimension order read reversed follwoing in matlab/fortran format
    '''

    #get file handle
    from netCDF4 import Dataset
    C=Dataset(fname,mode=mode);

    if fmt=='vars':
      svars=[*C.variables]; C.close(); return svars
    elif fmt==1:
        return ncfile(C,mode=mode) #return C
    elif fmt in [0,2]:
        F=zdata(); F.file_format=C.file_format

        #read dims
        ncdims=[i for i in C.dimensions];
        F.dimname=ncdims; F.dims=[]; F.dim_unlimited=[]
        for i in ncdims:
            F.dims.append(C.dimensions[i].size)
            F.dim_unlimited.append(C.dimensions[i].isunlimited())

        #read attrbutes
        ncattrs=list(setdiff1d(C.ncattrs(),['dims','dimname'])); F.attrs=ncattrs
        for i in ncattrs: F.__dict__[i]=C.getncattr(i)

        ncvars=[*C.variables]; F.vars=array(ncvars)
        #read variables
        for i in ncvars:
            if fmt==0:
               vi=zdata()
               dimi=C.variables[i].dimensions
               vi.dimname=dimi
               vi.dims=[C.dimensions[j].size for j in dimi]
               vi.val=C.variables[i][:]
               if hasattr(C.variables[i],'scale_factor'): vi.dtype=C.variables[i].dtype
               vi.attrs=C.variables[i].ncattrs()
               for j in C.variables[i].ncattrs():
                   ncattri=C.variables[i].getncattr(j)
                   vi.__dict__[j]=ncattri

               if order==1:
                   vi.dimname=list(flipud(vi.dimname))
                   vi.dims=list(flipud(vi.dims))
                   nm=flipud(arange(ndim(vi.val)))
                   vi.val=vi.val.transpose(nm)
            elif fmt==2:
               vi=array(C.variables[i][:],C.variables[i].dtype)
               if order==1: vi=vi.transpose(flip(arange(vi.ndim)))
            F.__dict__[i]=vi

        C.close()
        return F
    else:
        sys.exit('wrong fmt')

def WriteNC(fname,data,fmt=0,order=0,vars=None,**args):
    '''
    write zdata to netcdf file
        fname: file name
        data:  soure data (can be zdata or capsule of zdatas)
        fmt=0, data has zdata() format
        fmt=1, data has netcdf.Dataset format
        order=0: variable dimension order written not changed for python format
        order=1: variable dimension order written reversed follwoing in matlab/fortran format
        vars: list of variables that will be only saved
    '''
    from netCDF4 import Dataset

    if fmt==0: #write NC files for zdata
        #pre-processing
        if not fname.endswith('.nc'): fname=fname+'.nc'
        if not hasattr(data,'file_format'): data.file_format='NETCDF4'

        #check variables, dimensions, attrs
        sdict=data.__dict__; svars=[]; vdms=[]; vtypes=[]; vattrs=[]
        svars=[i for i,v in sdict.items() if (isinstance(v,zdata) or isinstance(v,np.ndarray)) and (i not in ['vars','dims'])] if (vars is None) else vars
        dims0=unique(concatenate([sdict[i].val.shape if hasattr(sdict[i],'val') else sdict[i].shape for i in svars]))
        if hasattr(data,'dims'):
           dims=list(data.dims); dnames=list(data.dimname)
           dtypes=list(data.dim_unlimited) if hasattr(data,'dim_unlimited') else [False for i in dims]
           if len(dtypes)<len(dims): dtypes=[*dtypes,*tile(False,len(dims)-len(dtypes))]
           for i in setdiff1d(dims0,dims): dims.append(i); dnames.append('dim_{}'.format(i)); dtypes.append(False)
        else:
           dims=dims0; dnames=['dim_{}'.format(i) for i in dims]; dtypes=[False for i in dims]; dims=list(dims)
        for i in svars:
          vdms.append(sdict[i].dimname if hasattr(sdict[i],'dimname') else [dnames[dims.index(k)] for k in sdict[i].shape])
          #vtypes.append(sdict[i].val.dtype if isinstance(sdict[i],zdata) else sdict[i].dtype)
          vtypes.append(sdict[i].dtype if hasattr(sdict[i],'dtype') else sdict[i].val.dtype)
          vattrs.append([k if k!='_FillValue' else '_fillvalue' for k in sdict[i].attrs] if hasattr(sdict[i],'attrs') else [])
        attrs=data.attrs if hasattr(data,'attrs') else setdiff1d([*sdict],[*svars,'vars','dims'])

        #write zdata as netcdf file
        fid=Dataset(fname,'w',format=data.file_format)   #open file
        fid.setncattr('file_format',data.file_format)    #set file_format
        for i in attrs: fid.setncattr(i,sdict[i])        #set global attrs
        for ds,dn,dt in zip(dims,dnames,dtypes): #set dimension
            fid.createDimension(dn,None) if (dt is True) else fid.createDimension(dn,ds)
        for i,svar in enumerate(svars):  #set variables
            vid=fid.createVariable(svar,vtypes[i],vdms[i] if order==0 else vdms[i][::-1],fill_value=sdict[svar]._FillValue if ('_fillvalue' in vattrs[i]) else None, **args)
            [vid.setncattr(k,sdict[svar].__dict__[k]) for k in vattrs[i] if k not in ['val','_fillvalue']]
            v=sdict[svar].val if isinstance(sdict[svar],zdata) else sdict[svar];
            fid.variables[svar][:]=v if order==0 else v.T
        fid.close()
    elif fmt==1: #write netcdf.Dateset as netcdf file
        C=data; cdict=C.__dict__; cdim=C.dimensions; cvar=C.variables
        fid=Dataset(fname,'w',format=C.file_format)     #open file
        fid.setncattr('file_format',C.file_format)      #set file_format
        for i in setdiff1d(C.ncattrs(),['dimname', 'dims', 'vars']): fid.setncattr(i,C.getncattr(i)) #set attrs
        if isinstance(C,ncfile): #set dims
           [fid.createDimension(i,None) if flag else fid.createDimension(i,k) for i,k,flag in zip(C.dimname,C.dims,C.dim_unlimited)]
        else:
           [fid.createDimension(i,None) if cdim[i].isunlimited() else fid.createDimension(i,cdim[i].size) for i in cdim]
        for i in cvar: #set vars
            if (vars is not None) and (i not in vars): continue
            vdim=cvar[i].dimensions
            vid=fid.createVariable(i,cvar[i].dtype,vdim) if order==0 else fid.createVariable(i,cvar[i].dtype,vdim[::-1])
            for k in cvar[i].ncattrs(): vid.setncattr(k,cvar[i].getncattr(k))
            if isinstance(C,ncfile):
               fid.variables[i][:]=cdict[i].value if order==0 else cdict[i].value.T
            else:
               fid.variables[i][:]=cvar[i][:] if order==0 else cvar[i][:].T
        fid.close()

def read(fname,*args0,**args):
    '''
    generic function in read files with standard suffixs
    suffix supported:  npz, pkl, gr3, ll, ic, vgrid.in, bp, reg, prop, xlsx, nc, shp, nml, asc, tif, tiff,mat
                       yaml,th,2dm
    for *.npz and *.nc file:
        read(fname,'IO'): return file channel
        read(fname,'vars'): return all variables
        read(fname,'INFO'): return all variables information
    '''
    from .schism_file import (read_schism_hgrid, read_schism_vgrid, read_schism_bpfile, read_schism_reg,
                              read_schism_prop, read_schism_param,read_schism_th,sms2grd)

    #determine read function
    if '.' not in fname: fname=fname+'.npz' #default file format
    F=None
    if fname.endswith('.asc') or fname.endswith('.tif') or fname.endswith('.tiff'): F=read_dem
    if fname.endswith('.gr3') or fname.endswith('.ll') or fname.endswith('.ic'): F=read_schism_hgrid
    if fname.endswith('vgrid.in'):  F=read_schism_vgrid
    if fname.endswith('.npz') or fname.endswith('pkl') or fname.endswith('.pp'):  F=loadz
    if fname.endswith('.bp'):   F=read_schism_bpfile
    if fname.endswith('.reg'):  F=read_schism_reg
    if fname.endswith('.prop'): F=read_schism_prop
    if fname.endswith('.xlsx'): F=read_excel
    if fname.endswith('.nc'):   F=ReadNC
    if fname.endswith('.yaml'): F=read_yaml
    if fname.endswith('.th'):   F=read_schism_th
    if fname.endswith('.shp'):  F=read_shapefile_data
    if fname.endswith('.2dm'):  F=sms2grd
    if fname.endswith('.nml'):  F=read_schism_param
    if fname.endswith('.mat'):  F=read_mat
    if F is None: sys.exit('unknown type of file: '+fname)

    if fname.endswith('.npz') or fname.endswith('.nc'):
       def fid_npz(svar):
           return loadz(fname,svar)
       def fid_nc(svar):
           fid=ReadNC(fname,1); f=fid.variables; data=[*f] if svar=='vars' else array(f[svar][:]); fid.close()
           return data
       fid=fid_npz if fname.endswith('.npz') else fid_nc

       #return information about npz or netcdf file
       if 'IO' in args0:
          return fid
       elif 'vars' in args0:
          return fid('vars')
       elif ('INFO' in args0) or ('VINFO' in args0):
          vs=fid('vars'); ms=min([6,max([len(i) for i in vs])]); fs0='{:'+str(ms)+'s}: '
          lines=[]; dt=''; ds=''
          for i in vs:
              if i.startswith('_') and i.endswith('_variables'): continue
              v=fid(i); t=type(v); line=''
              if t is list: line=fs0.format(i)+'list({})'.format(len(v))
              if t is dict: line=fs0.format(i)+'dict({})'.format(len(v))
              if t is str:  line=fs0.format(i)+v[:30]+',string'
              if t is np.ndarray: line=fs0.format(i)+'array{},{}'.format(v.shape,v.dtype)
              if line=='': line=fs0.format(i)+str(t)
              lines.append(line)
          return lines
       else:
          return F(fname,*args0,**args)
    else:
        return F(fname,*args0,**args)

def pplot(fnames):
    '''
    function to display figures in python format (*.pp)
    '''
    if isinstance(fnames,str): fnames=[fnames]
    try:
        hfs=[]
        for fname in fnames:
            F=read(fname)
            if hasattr(F,'attrs'): #redraw
               hf=figure()
               for ax0 in F.axs:
                   ax=axes(ax0.position)
                   for hp0 in ax0.ps:
                       if hp0.type=='line':
                           hp=plot(hp0.x,hp0.y)[0]
                       elif hp0.type=='text':
                           hp=text(*hp0.postion,hp0.text)
                       elif hp0.type=='legend':
                           hp=legend(hp0.texts); hl=hp; pos=hp0.position
                           hp.set_bbox_to_anchor(hp0.position,transform=hf.transFigure)
                       for i in hp0.attrs: exec('hp.set_{}(hp0._{})'.format(i,i))
                   for i in ax0.attrs: exec('ax.set_{}(ax0._{})'.format(i,i))
                   for i,k in zip(ax.get_xgridlines(),ax0.xgrid): i.set_visible(k)
                   for i,k in zip(ax.get_ygridlines(),ax0.ygrid): i.set_visible(k)
               for i in F.attrs: exec('hf.set_{}(F._{})'.format(i,i))
               #hl.set_bbox_to_anchor(pos,transform=hf.transFigure)
            else:
               hf=F.hf
        hfs.append(hf)
    except:
        import pickle
        hfs=[pickle.load(open(i,'rb')) for i in fnames]

    show(block=False)
    return hfs

def savefig(fname,fig=None,fmt=0,**args):
    '''
    rewrite python savefig function with new options
    fname: figure name
           if fname.endswith('.pp'):
              fmt==0: extract data from figure, to replot later
              fmt==1: save fig in binary format (can repeat precisely, but with large memory)
           else:
              plt.savefig(fname)
    '''
    if isinstance(fname,str) and fname.endswith('.pp'):
       hf=gcf() if fig is None else fig
       if fmt==0:
          #fig info
          attrs=['figheight','figwidth','tight_layout','zorder']
          F=zdata(); F.attrs=attrs; F.axs=[]
          for i in attrs: exec('F._{}=hf.get_{}()'.format(i,i))

          for ax in hf.axes: #for axes
              #save ax info
              attrs=['aspect','facecolor','label','title','xlabel','xscale','xticks','xticklabels',
                     'ylabel','yscale','yticks','yticklabels','xlim','ylim','zorder',]
              A=zdata(); A.attrs=attrs; A.ps=[]; A.position=ax.get_position()
              for i in attrs: exec('A._{}=ax.get_{}()'.format(i,i))
              A._xticklabels=[i.get_text() for i in A._xticklabels]
              A._yticklabels=[i.get_text() for i in A._yticklabels]
              A.xgrid=[i.get_visible() for i in ax.get_xgridlines()]
              A.ygrid=[i.get_visible() for i in ax.get_ygridlines()]

              for hp in ax.get_children(): #save info for each plot obj
                  if type(hp)==mpl.lines.Line2D:
                      attrs=['color','linestyle','linewidth','marker','markeredgecolor','markeredgewidth',
                             'markerfacecolor','markerfacecoloralt','markersize','markevery','zorder',]
                      H=zdata(); H.type='line'; H.attrs=attrs; H.x,H.y=hp.get_xdata(), hp.get_ydata()
                      for i in attrs: exec('H._{}=hp.get_{}()'.format(i,i))
                  elif type(hp)==mpl.text.Text:
                      attrs=['color','fontsize','fontweight','rotation']
                      H=zdata(); H.type='text'; H.attrs=attrs; H.postion,H.text=hp.get_position(),hp.get_text()
                      for i in attrs: exec('H._{}=hp.get_{}()'.format(i,i))
                  elif type(hp)==mpl.legend.Legend:
                      attrs=[]
                      H=zdata(); H.type='legend'; H.attrs=attrs
                      H.texts=[i.get_text() for i in hp.get_texts()]
                      for i in attrs: exec('H._{}=hp.get_{}()'.format(i,i))
                      pl=hp.get_bbox_to_anchor(); pf=hf.bbox; xm=pf.x1-pf.x0; ym=pf.y1-pf.y0
                      H.position=[pl.x0/xm,pl.y0/ym,(pl.x1-pl.x0)/xm,(pl.y1-pl.y0)/ym,]
                  else:
                      continue
                  A.ps.append(H)
              F.axs.append(A)
          F.save(fname)
       else: #fmt==1: binary format
          try:
             c=zdata(); c.hf=hf; c.save(fname)
          except:
             import pickle
             pickle.dump(hf,open(fname, "wb"))
    else:
       plt.savefig(fname,**args)

def harmonic_fit(oti,oyi,dt,mti,tidal_names=0, **args):
    '''
    construct time series using harmonics: used to fill data gap, or extrapolation
       [oti,oyi,dt]: observed (time, value, time interval)
       mti: time to interpolate/extrapolate
       tidal_names=0/1/2: tidal names options; tidal_names=['O1','K1','Q1','P1','M2','S2',]: list of tidal consituent names
    '''
    #choose tidal consts
    if not hasattr(tidal_names,'__len__'):
       if tidal_names==0:
          tidal_names=array(['O1','K1','Q1','P1','M2','S2','K2','N2'])
       elif tidal_names==1:
          tidal_names=array(['O1','K1','Q1','P1','M2','S2','K2','N2','M3','M4','M6','M7','M8','M10'])
       elif tidal_names==2:
          tidal_names=array(['SA','SSA','MM','MSF','O1','K1','Q1','P1','M2','S2','K2','N2','M3','M4','M6','M7','M8','M10','N4','S4','S6'])

    #do harmnoic analysis
    H=harmonic_analysis(oyi,dt,tidal_names=tidal_names,**args)

    #fit data
    myi=zeros(mti.shape)+H.amplitude[0]
    for k,tname in enumerate(tidal_names): myi=myi+H.amplitude[k+1]*cos(H.freq[k+1]*(mti-oti[0])*86400-H.phase[k+1])

    return myi

def harmonic_analysis(data,dt,t0=0,tidal_names=None,code=None,tname=None,fname=None,sname=None):
    '''
    harmonic analyze time series
        data: time series
        dt: time step (day)
        t0: starting time of time series (day); normally use datenum (e.g. datenum(2010,0,0))
        tidal_names: 1) path of tidal_const.dat, or 2) names of tidal consituents
        code: path of executable (tidal_analyze, or tidal_analyze.exe) for HA analysis
        [tname,fname,sname]: temporary names for tidal_const, time series, and HA results
    '''
    import subprocess

    #names of temporary file
    if tname is None: tname='.temporary_tidal_const_for_HA.dat'
    if fname is None: fname='.temporary_time_series_for_HA.dat'
    if sname is None: sname='.temporary_tidal_consituents_for_HA.dat'

    #check OS type, and locate the executable
    sdir='{}/../scripts/Harmonic_Analysis'.format(os.path.dirname(__file__))
    if code is None:
        import platform
        #directories where exectuable may exist
        bdirs=[sdir, r'D:\Work\Harmonic_Analysis',r'~/bin/Harmonic_Analysis']
        if platform.system().lower()=='windows':
           code=['{}/tidal_analyze.exe'.format(i) for i in bdirs if os.path.exists('{}/tidal_analyze.exe'.format(i))][0]
        elif platform.system().lower()=='linux':
           code=['{}/tidal_analyze'.format(i) for i in bdirs if os.path.exists('{}/tidal_analyze'.format(i))][0]
        else:
            print('Operating System unknow: {}'.format(platform.system()))
        if code is None: sys.exit('exectuable "tidal_analyze" was not found')

    #locate or write tidal_const.dat
    if tidal_names is None:
       tidal_names='{}/tidal_const.dat'.format(sdir)
    else:
       if not isinstance(tidal_names,str): #names of tidal consituents
          C=loadz('{}/tide_fac_const.npz'.format(sdir)); tdict=dict(zip(C.name,C.freq))
          fid=open(tname,'w+'); fid.write('{}\n'.format(len(tidal_names)))
          for i in tidal_names: fid.write('{}\n{}\n'.format(i.upper(),tdict[i.upper()]))
          fid.close(); tidal_names=tname

    #write time series, HA, and return results
    fid=open(fname,'w+')
    for i,datai in enumerate(data): fid.write('{:12.1f} {:12.7f}\n'.format((t0+i*dt)*86400,datai))
    fid.close(); subprocess.call([code,fname,tidal_names,sname]) #HA

    #save results
    S=zdata(); S.tidal_name,S.amplitude,S.phase=array([i.strip().split() for i in open(sname,'r').readlines()]).T
    S.amplitude=array(S.amplitude).astype('float'); S.phase=array(S.phase).astype('float')
    S.freq=r_[0,array([i.strip() for i in open(tidal_names,'r').readlines()[2::2]]).astype('float')]

    #clean temporaray files--------
    os.remove(fname); os.remove(sname)
    if os.path.exists(tname): os.remove(tname)
    return S

def ceqstate(T,S,P=0):
    '''
    calculate seawater density (Millero and Poisson 1981, Gill, 1982); Input:
      T: temperature in [C]
      S: salinity in [PSU]
      P: pressure [bars] (P=0: at 1 atm. pressure)
    '''

    npt=len(T)
    #pre_calculation
    T2=T*T; T3=T**3; T4=T**4; T5=T**5
    S05=sqrt(S); S15=S*S05; S2=S*S; P2=P*P

    #pure water S=0,at 1 atm.
    rho_pw=999.842594+6.793952e-2*T-9.095290e-3*T2+1.001685e-4*T3-1.120083e-6*T4+6.536332e-9*T5

    #density with Salinity
    A=8.24493e-1-4.0899e-3*T+7.6438e-5*T2-8.2467e-7*T3+5.3875e-9*T4
    B=-5.72466e-3+1.0227e-4*T-1.6546e-6*T2
    C=4.8314e-4
    rho_st=rho_pw+A*S+B*S15+C*S2

    #pressure not zero
    if P==0:
      rho=rho_st
    else:
      K_pw=19652.21+148.4206*T-2.327105*T2+1.360477e-2*T3-5.155288e-5*T4; K_st=K_pw
      K_st=K_st+S*(54.6746-0.603459*T+1.09987e-2*T2-6.1670e-5*T3)+S15*(7.944e-2+1.6483e-2*T-5.3009e-4*T2);
      K_stp=K_st+P*(3.239908+1.43713e-3*T+1.16092e-4*T2-5.77905e-7*T3)+P*S*(2.2838e-3-1.0981e-5*T-1.6078e-6*T2)+1.91075e-4*P*S15 \
            +P2*(8.50935e-5-6.12293e-6*T+5.2787e-8*T2)+P2*S*(-9.9348e-7+2.0816e-8*T+9.1697e-10*T2)
      rho=rho_st/(1-P/K_stp)
    return rho

def subdomain_index(x,y,xm,ym):
    '''
    get subdomain indice (e.g. for subsetting purpose)
    Inputs:
       x(ny,nx): x coordinate
       y(ny,nx): y coordinate
       xm:  x-range: [-75, -70],  or x-array (eg. gd.x)
       ym:  y-range: [35, 40],    or y-array (eg. gd.y)
    Output:  ix1,ix2,iy1,iy2
    '''
    #pre-proc
    if len(xm)==2: xm=array([*xm,xm[1],xm[0]]); ym=array([ym[0],ym[0],ym[1],ym[1]])
    xc=xm.mean(); yc=xm.mean(); x1,x2,y1,y2=xm.min(),xm.max(),ym.min(),ym.max()
    ny,nx=x.shape; sind=near_pts(c_[xc,yc],c_[x.ravel(),y.ravel()])
    iy,ix=[i[0] for i in unravel_index(sind,[ny,nx])]; ix1,ix2,iy1,iy2=ix,ix+1,iy,iy+1

    #find initial box
    while True:
        ix1i=ix1; ix2i=ix2; iy1i=iy1; iy2i=iy2
        if x[iy1:iy2,ix1].max()>x1: ix1=max([0,ix1-1])
        if x[iy1:iy2,ix2-1].min()<x2: ix2=min([nx,ix2+1])
        if y[iy1,ix1:ix2].max()>y1: iy1=max([0,iy1-1])
        if y[iy2-1,ix1:ix2].min()<y2: iy2=min([ny,iy2+1])
        if (ix1i==ix1)*(ix2i==ix2)*(iy1i==iy1)*(iy2i==iy2): break

    #shrink the box
    for m in arange(2):
        if m==1 and len(xm)==2: continue
        pxy=c_[array([x1,x2,x2,x1]),array([y1,y1,y2,y2])] if m==0 else c_[xm,ym]
        for n in arange(4):
            while True:
                if n==0: ix1=ix1+1
                if n==1: ix2=ix2-1
                if n==2: iy1=iy1+1
                if n==3: iy2=iy2-1
                xs=r_[x[iy1:iy2,ix1],x[iy2-1,(ix1+1):(ix2-1)],x[iy1:iy2,ix2-1][::-1],x[iy1,(ix1+1):(ix2-1)][::-1]]
                ys=r_[y[iy1:iy2,ix1],y[iy2-1,(ix1+1):(ix2-1)],y[iy1:iy2,ix2-1][::-1],y[iy1,(ix1+1):(ix2-1)][::-1]]
                if inside_polygon(pxy,xs,ys).min()==0:
                   if n==0: ix1=ix1-1
                   if n==1: ix2=ix2+1
                   if n==2: iy1=iy1-1
                   if n==3: iy2=iy2+1
                   break
    return ix1,ix2,iy1,iy2

def get_hycom(Time,xyz,vind,hdir='./HYCOM',method=0):
    '''
    extract Hycom time series at stations
    ti: time seires; xyz=c_[loni,lati,depi];
    vind: list of index for variables to be extracted. [0,1,2,3] for ['elev','temp','salt','uv']
    hdir: directory for hycom data
    method=0: linear interpolation; method=1: nearest interpolation
    '''

    #variable names
    Var=['surf_el','water_temp','salinity',['water_u','water_v']];
    VarName=['elev','temp','salt',['Ux','Uy']];

    #time
    StartT=floor(Time.min())
    EndT=ceil(Time.max())

    #---interpolation pts----
    if xyz.ndim==1: xyz=xyz[None,:]
    loni=xyz[:,0]; lati=xyz[:,1]
    bxy=c_[lati,loni];
    if xyz.shape[1]==3:
       depi=xyz[:,2]
       bxyz=c_[depi,lati,loni];

    #-----save data----------------
    S=zdata();
    S.x=xyz[:,0]; S.y=xyz[:,1]
    if xyz.shape[1]==3: S.z=xyz[:,2]

    #---read Hycom data and interpolate onto boundary nodes------------------
    p=zdata();
    for i in vind: #arange(len(Var)):
        vari=Var[i]; varnamei=VarName[i];

        t0=time.time();
        T0=[]; Data0=[];
        for ti in arange(StartT,EndT+1):
            t1=num2date(ti); t2=num2date(ti+1-1/24/60);

            if isinstance(vari,list):
                fname='Hycom_{}_{}_{}.nc'.format(varnamei[0],t1.strftime('%Y%m%dZ%H%M00'),t2.strftime('%Y%m%dZ%H%M00'))
                fname2='Hycom_{}_{}_{}.nc'.format(varnamei[1],t1.strftime('%Y%m%dZ%H%M00'),t2.strftime('%Y%m%dZ%H%M00'))
                if not os.path.exists(r'{}/{}'.format(hdir,fname)): continue
                if not os.path.exists(r'{}/{}'.format(hdir,fname2)): continue
                print(fname+'; '+fname2)
                C=ReadNC('{}/{}'.format(hdir,fname)); C2=ReadNC('{}/{}'.format(hdir,fname2),2)

                #get value
                exec('p.val=C.{}.val'.format(vari[0]))
                exec('p.val2=C2.{}.val'.format(vari[1]))
                p.val=array(p.val); fp=p.val<-29999; p.val2=array(p.val2); fp2=p.val2<-29999;
                p.val[fp]=nan; p.val2[fp2]=nan;
            else:
                fname='Hycom_{}_{}_{}.nc'.format(varnamei,t1.strftime('%Y%m%dZ%H%M00'),t2.strftime('%Y%m%dZ%H%M00'))
                if not os.path.exists(r'{}/{}'.format(hdir,fname)): continue
                print(fname)
                C=ReadNC('{}/{}'.format(hdir,fname))

                #get value
                exec('p.val=C.{}.val'.format(vari))
                p.val=array(p.val); fp=p.val<-29999;
                p.val[fp]=nan

            ti=datestr2num(C.time.time_origin)+array(C.time.val)/24
            cloni=array(C.lon.val); cloni=mod(cloni,360)-360
            clati=array(C.lat.val);

            #------define data region extracted
            ind_lon=pindex((cloni<=max(loni)+0.1)*(cloni>=min(loni)-0.1))
            ind_lat=pindex((clati<=max(lati)+0.1)*(clati>=min(lati)-0.1))
            i1_lon=ind_lon.min(); i2_lon=i1_lon+len(ind_lon)
            i1_lat=ind_lat.min(); i2_lat=i1_lat+len(ind_lat)

            cloni=cloni[i1_lon:i2_lon]; clati=clati[i1_lat:i2_lat]

            if varnamei=='elev':
                for m in arange(len(ti)):
                    valii=squeeze(p.val[m,i1_lat:i2_lat,i1_lon:i2_lon])

                    if method==0:
                        #interpolation
                        fd=sp.interpolate.RegularGridInterpolator((clati,cloni),valii,fill_value=nan)
                        vi=fd(bxy)

                        #remove nan pts
                        fp=isnan(vi);
                        if sum(fp)!=0:
                            vi[fp]=sp.interpolate.griddata(bxy[~fp,:],vi[~fp],bxy[fp,:],'nearest')

                    elif method==1:
                        clonii,clatii=meshgrid(cloni,clati); sind=~isnan(valii.ravel())
                        clonii=clonii.ravel()[sind]; clatii=clatii.ravel()[sind]; valii=valii.ravel()[sind]
                        vi=sp.interpolate.griddata(c_[clatii,clonii],valii,bxy,'nearest')
                        # scatter(clonii,clatii,s=8,c=valii); plot(bxy[:,1],bxy[:,0],'r.'); sys.exit()
                    T0.append(ti[m]); Data0.append(vi);
            else:
                #------define data region extracted for depth
                cdepi=array(C.depth.val)
                ind_dep=pindex((cdepi<=depi.max()+1000)*(cdepi>=depi.min()-100))
                i1_dep=ind_dep.min(); i2_dep=i1_dep+len(ind_dep)
                cdepi=cdepi[i1_dep:i2_dep];

                for m in arange(len(ti)):
                    T0.append(ti[m]);
                    valii=squeeze(p.val[m,i1_dep:i2_dep,i1_lat:i2_lat,i1_lon:i2_lon])

                    #interpolation
                    fd=sp.interpolate.RegularGridInterpolator((cdepi,clati,cloni),valii,fill_value=nan)
                    vi=fd(bxyz)

                    #remove nan pts
                    fp=isnan(vi);
                    if sum(fp)!=0:
                        vi[fp]=sp.interpolate.griddata(bxyz[~fp,:],vi[~fp],bxyz[fp,:],'nearest')

                    #----if variable is velocity
                    if isinstance(varnamei,list):
                        val2ii=squeeze(p.val2[m,i1_dep:i2_dep,i1_lat:i2_lat,i1_lon:i2_lon])

                        #interpolation
                        fd=sp.interpolate.RegularGridInterpolator((cdepi,clati,cloni),val2ii,fill_value=nan)
                        v2i=fd(bxyz)

                        #remove nan pts
                        fp=isnan(v2i);
                        if sum(fp)!=0:
                            v2i[fp]=sp.interpolate.griddata(bxyz[~fp,:],v2i[~fp],bxyz[fp,:],'nearest')

                        Data0.append(r_[expand_dims(vi,0),expand_dims(v2i,0)])

                    else:
                        Data0.append(vi)

        #interpolate in time, and save data
        T0=array(T0); Data0=array(Data0).T; S.time=Time;
        if Data0.ndim==2:
            #interpolate
            datai=[];
            for k in arange(Data0.shape[0]):
                fd=sp.interpolate.interp1d(T0,Data0[k],fill_value='extrapolate');
                datai.append(fd(Time));
            datai=array(datai)
            exec('S.{}=datai'.format(varnamei))
        else:
            #interpolate
            datai=[]; data2i=[];
            for k in arange(Data0.shape[0]):
                fd=sp.interpolate.interp1d(T0,Data0[k,0]);
                fd2=sp.interpolate.interp1d(T0,Data0[k,1]);
                datai.append(fd(Time)); data2i.append(fd2(Time))
            datai=array(datai); data2i=array(data2i)
            exec('S.{}=datai'.format(varnamei[0]))
            exec('S.{}=data2i'.format(varnamei[1]))

    return S

if __name__=="__main__":
    pass
